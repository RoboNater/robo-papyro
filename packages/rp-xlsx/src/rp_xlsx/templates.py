"""Template resolution, inspection, manifests, and synthesis (spec section 5).

**Resolution follows `rp-docx`'s convention, not `rp-pptx`'s divergent one**
(spec section 5.1): ``RP_XLSX_TEMPLATE_DIR`` splits on ``os.pathsep`` and
searches ancestor repo roots, matching ``RP_DOCX_TEMPLATE_DIR``.
``RP_PPTX_TEMPLATE_DIR`` diverged (a single directory, ``./templates`` only)
and ``AGENTS.md`` records that as a written-up gap rather than a pattern to
repeat.

**``resolve_template(None)`` returns ``None`` here, unlike its two siblings**
(spec section 4). openpyxl ships no bundled workbook to fall back to —
``Workbook()`` is not a template, it is an empty file — so inventing a
default would mean shipping a binary or synthesizing one nobody asked for.
``create(template=None)`` starts from ``openpyxl.Workbook()`` directly.

**The manifest loop exists because real templates cannot enter this
repository** (spec section 5.2). ``build_manifest`` runs against the real
file wherever it lives and emits JSON that is redacted by construction — no
document text besides the header row and placeholder cells, which are
structure rather than content; ``synthesize`` reconstructs a structurally
equivalent template at test time. ``SheetShape`` is the redacted shape; see
its docstring in ``models.py`` for exactly what does and does not survive.

Nothing here prints, and nothing here imports typer.
"""

from __future__ import annotations

import io
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from rp_xlsx import ooxml, refs
from rp_xlsx.errors import TemplateError
from rp_xlsx.models import SheetShape, TemplateInfo, TemplateManifest
from rp_xlsx.ooxml import SUPPORTED_SUFFIXES, TEMPLATE_SUFFIXES

#: ``{{ key }}`` / ``{{ key.subkey }}`` with any amount of inner whitespace.
#: Keys are dotted identifiers only — no calls, no indexing, no operators, so
#: there is no expression here to evaluate even by accident (spec section 8).
PLACEHOLDER = re.compile(r"\{\{\s*([A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)*)\s*\}\}")

TEMPLATE_DIR_ENV = "RP_XLSX_TEMPLATE_DIR"
DEFAULT_TEMPLATE_ENV = "RP_XLSX_TEMPLATE"

#: `.xltx` before `.xltm` before `.xlsx`, matching the order spec section 5.1
#: gives: a template is what was asked for, and a macro-enabled one is more
#: specific than a plain workbook of the same name.
LOOKUP_SUFFIXES = (*TEMPLATE_SUFFIXES, ".xlsx")

assert set(LOOKUP_SUFFIXES) <= set(SUPPORTED_SUFFIXES)  # keep in sync with ooxml.py


def repo_root(start: Path | None = None) -> Path | None:
    """The nearest ancestor of ``start`` that looks like the project checkout.

    Templates live in ``<repo>/templates/``, which only means anything when
    running from a checkout — an installed wheel has no repo. The marker is a
    ``templates`` directory next to a ``.git`` or ``pyproject.toml``, so an
    unrelated ``templates/`` in some working directory is not mistaken for the
    suite's.
    """
    current = (start or Path.cwd()).resolve()
    for candidate in (current, *current.parents):
        if (candidate / "templates").is_dir() and (
            (candidate / ".git").exists() or (candidate / "pyproject.toml").is_file()
        ):
            return candidate
    return None


def template_dirs() -> list[Path]:
    """Where a bare template name is looked up, in precedence order.

    ``RP_XLSX_TEMPLATE_DIR`` first (it may name several directories, split
    the way ``PATH`` is), then the checkout's ``templates/local/`` and
    ``templates/``. ``local/`` comes first because it is the gitignored drop
    point for the *real* templates (spec section 11.1): when a name exists in
    both, the real one is the one meant.
    """
    dirs: list[Path] = []
    configured = os.environ.get(TEMPLATE_DIR_ENV, "")
    for entry in configured.split(os.pathsep):
        if entry.strip():
            dirs.append(Path(entry.strip()).expanduser())
    root = repo_root()
    if root is not None:
        dirs.extend([root / "templates" / "local", root / "templates"])
    return [d for d in dirs if d.is_dir()]


def _lookup(name: str) -> Path | None:
    for directory in template_dirs():
        for suffix in LOOKUP_SUFFIXES:
            candidate = directory / f"{name}{suffix}"
            if candidate.is_file():
                return candidate
    return None


def available_template_names() -> list[str]:
    """Bare names :func:`resolve_template` would find, de-duplicated."""
    names: list[str] = []
    for directory in template_dirs():
        for suffix in LOOKUP_SUFFIXES:
            names.extend(path.stem for path in sorted(directory.glob(f"*{suffix}")))
    return sorted(dict.fromkeys(names))


def _looks_like_a_path(value: str) -> bool:
    """Whether the argument is a path the user got wrong, or a name to look up.

    Anything carrying a suffix or a separator was meant as a path (spec
    section 5.1 case 4), and reporting "no template called
    ../drafts/quarterly.xltx" would send the user hunting through the
    template directories for a typo in their own path.
    """
    as_path = Path(value)
    return bool(as_path.suffix) or os.sep in value or (os.altsep and os.altsep in value)


def resolve_template(name_or_path: str | Path | None = None) -> Path | None:
    """Find the template a caller means (spec section 5.1).

    1. An existing path is used as given
    2. A bare name resolves against :func:`template_dirs`, `.xltx` before
       `.xltm` before `.xlsx`
    3. ``None`` uses ``RP_XLSX_TEMPLATE`` if set, else ``None`` — this
       package's considered divergence from its two siblings (module
       docstring)
    4. A path-shaped argument that does not exist raises, naming the *path*
    5. An unresolvable name raises, listing the available templates
    """
    if name_or_path is None:
        configured = os.environ.get(DEFAULT_TEMPLATE_ENV, "").strip()
        return resolve_template(configured) if configured else None

    as_path = Path(name_or_path)
    if as_path.is_file():
        return as_path

    text = str(name_or_path)
    if _looks_like_a_path(text):
        raise TemplateError(f"No such template file: {text}")

    found = _lookup(text)
    if found is not None:
        return found

    names = available_template_names()
    listed = ", ".join(names) if names else "none found"
    raise TemplateError(f"Unknown template {text!r}; available templates: {listed}")


# --- inspection --------------------------------------------------------------


def list_templates() -> list[TemplateInfo]:
    """Every resolvable template, inspected. First directory wins on a name clash."""
    found: dict[str, Path] = {}
    for directory in template_dirs():
        for suffix in LOOKUP_SUFFIXES:
            for path in sorted(directory.glob(f"*{suffix}")):
                found.setdefault(path.stem, path)
    return [inspect_template(path) for _, path in sorted(found.items())]


def placeholder_for(key: str) -> str:
    return "{{ " + key + " }}"


def find_placeholders(path: Path) -> list[str]:
    """Every distinct ``{{ key }}`` in the workbook's cell values and
    header/footer text, in first-seen order (sheet order, then row-major).

    Walks the same surfaces :func:`~rp_xlsx.xlsx.write.replace_text` does, so
    a placeholder hiding in a header is found before someone discovers it
    only once printed.
    """
    seen: list[str] = []
    with ooxml.opened(path) as wb:
        for ws in wb.worksheets:
            for cell in ooxml.populated_cells(ws):
                if isinstance(cell.value, str):
                    for key in PLACEHOLDER.findall(cell.value):
                        if key not in seen:
                            seen.append(key)
            for _label, part in ooxml.header_footer_fields(ws):
                if part.text:
                    for key in PLACEHOLDER.findall(part.text):
                        if key not in seen:
                            seen.append(key)
    return seen


def inspect_template(path: Path) -> TemplateInfo:
    """A template's shape: sheets, defined names, and the placeholders it
    carries (spec section 4)."""
    from rp_xlsx.xlsx.read import get_index, get_names

    index = get_index(path)
    return TemplateInfo(
        name=Path(path).stem,
        path=Path(path),
        format=index.format,
        sheets=index.sheets,
        defined_names=get_names(path),
        placeholders=find_placeholders(path),
    )


# --- manifests and synthesis (spec section 5.2) ------------------------------


def _cell_text(cell: Any) -> str:
    """A header cell's text, normalizing a whitespace-only value to "".

    openpyxl does not round-trip a truly empty string (verified: it comes
    back ``None`` after a save/reload, unlike every other value including a
    bare space) — so treating "just whitespace" the same as "empty" here is
    what lets a header built with :func:`synthesize` compare equal to the
    original after a reload, and it costs nothing a manifest cares about:
    exact whitespace inside a header cell is not structure.
    """
    if cell.value is None:
        return ""
    text = str(cell.value)
    return "" if not text.strip() else text


def _uniform_number_formats(ws: Any, bounds: tuple[int, int, int, int]) -> dict[str, str]:
    """Column letter -> number format, only where every populated cell in the
    column (excluding the header row) shares one format. A format string is
    presentation and the thing most likely to be wrong, so a column that
    disagrees with itself contributes nothing rather than a guess."""
    min_row, min_col, max_row, max_col = bounds
    formats: dict[str, str] = {}
    for col in range(min_col, max_col + 1):
        seen: set[str] = set()
        for row in range(min_row + 1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                seen.add(cell.number_format)
        if len(seen) == 1:
            formats[refs.column_letters(col)] = next(iter(seen))
    return formats


def _placeholder_cells(ws: Any, bounds: tuple[int, int, int, int]) -> dict[str, str]:
    min_row, min_col, max_row, max_col = bounds
    found: dict[str, str] = {}
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and PLACEHOLDER.search(cell.value):
                found[f"{refs.column_letters(col)}{row}"] = cell.value
    return found


def _sheet_shape(ws: Any, index: int) -> SheetShape:
    from rp_xlsx.xlsx.read import _used_bounds

    bounds = _used_bounds(ws)
    header: list[str] | None = None
    used_range: str | None = None
    number_formats: dict[str, str] = {}
    placeholder_cells: dict[str, str] = {}
    if bounds is not None:
        min_row, min_col, max_row, max_col = bounds
        used_range = (
            f"{refs.column_letters(min_col)}{min_row}:{refs.column_letters(max_col)}{max_row}"
        )
        header = [_cell_text(ws.cell(row=min_row, column=c)) for c in range(min_col, max_col + 1)]
        number_formats = _uniform_number_formats(ws, bounds)
        placeholder_cells = _placeholder_cells(ws, bounds)
    column_widths = {
        letter: dim.width for letter, dim in ws.column_dimensions.items() if dim.width is not None
    }
    table_names = list(ws.tables)
    return SheetShape(
        index=index,
        name=ws.title,
        state=ws.sheet_state,
        used_range=used_range,
        header=header,
        column_widths=column_widths,
        freeze_panes=ws.freeze_panes,
        number_formats=number_formats,
        table_names=table_names,
        placeholder_cells=placeholder_cells,
    )


def build_manifest(path: Path) -> TemplateManifest:
    """A redacted-by-construction description of ``path``'s shape (spec
    section 5.2). No cell value enters it except a sheet's header row and its
    placeholder cells, both of which are structure rather than content —
    ``tests/test_templates_xlsx.py`` asserts that a manifest built from a
    template with distinctive *body* text does not contain that text
    anywhere in its serialized form.
    """
    from rp_xlsx.xlsx.read import get_names

    with ooxml.opened(path) as wb:
        sheets = [_sheet_shape(ws, i) for i, ws in enumerate(wb.worksheets, start=1)]
        image_count = sum(len(ws._images) for ws in wb.worksheets)
    return TemplateManifest(
        name=Path(path).stem,
        format=ooxml.format_of(Path(path)),
        sheets=sheets,
        defined_names=get_names(path),
        placeholders=find_placeholders(path),
        image_count=image_count,
    )


#: A 1x1 transparent-ish PNG, standing in for "a logo is here" (spec section
#: 5.2: "logo presence, not logo bytes"). Committed as bytes, not a binary
#: file — spec section 11.1 forbids the latter.
_PLACEHOLDER_LOGO_PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108020000009077"
    "53de0000000c4944415408d763f8ffff3f0005fe02fea739669d0000000049"
    "454e44ae426082"
)


def synthesize(manifest: TemplateManifest, output: Path) -> Path:
    """Reconstruct a structurally equivalent template from a manifest.

    Reproduces sheet names, order, and visibility; header rows; column
    widths; frozen panes; per-column number formats; defined names; table
    names (over a synthetic ref covering the header and one data row —
    the manifest carries no column definitions to reproduce, only names);
    placeholder cells; and a placeholder image when ``image_count > 0``.
    Does **not** reproduce themes, fonts, colours, conditional formatting, or
    data validation — structural equivalence for testing resolution and
    filling, not visual fidelity (spec section 5.2).
    """
    output = Path(output)
    wb = Workbook()
    default_name = wb.sheetnames[0]

    for shape in manifest.sheets:
        ws = wb.create_sheet(shape.name)
        ws.sheet_state = shape.state
        if shape.header:
            for col, text in enumerate(shape.header, start=1):
                ws.cell(row=1, column=col, value=text)
        for ref, text in shape.placeholder_cells.items():
            position = refs.parse_cell_ref(ref)
            ws.cell(row=position.row, column=position.column, value=text)
        for letter, width in shape.column_widths.items():
            ws.column_dimensions[letter].width = width
        data_row = 2 if shape.header else 1
        for letter, number_format in shape.number_formats.items():
            cell = ws.cell(row=data_row, column=refs.column_index(letter))
            # A format with no value is invisible to a reread: the reader
            # only notices number_format on a cell that also holds a value
            # (verified -- a value-less formatted cell is skipped as "not
            # used"), so a representative value is required here too. Its
            # type does not need to match the format -- structural
            # equivalence, not visual fidelity (spec section 5.2).
            if cell.value is None:
                cell.value = 0
            cell.number_format = number_format
        if shape.freeze_panes:
            ws.freeze_panes = shape.freeze_panes
        if shape.header:
            last_col = refs.column_letters(len(shape.header))
            table_ref = f"A1:{last_col}{data_row}"
            for name in shape.table_names:
                table = Table(displayName=name, ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
                ws.add_table(table)
        if manifest.image_count and shape.index == 1:
            ws.add_image(Image(io.BytesIO(_PLACEHOLDER_LOGO_PNG)), "A1")

    del wb[default_name]

    for name in manifest.defined_names:
        target = wb if name.scope is None else wb[name.scope]
        target.defined_names[name.name] = DefinedName(name.name, attr_text=name.refers_to)

    return ooxml.save(wb, output)


__all__ = [
    "DEFAULT_TEMPLATE_ENV",
    "LOOKUP_SUFFIXES",
    "PLACEHOLDER",
    "TEMPLATE_DIR_ENV",
    "available_template_names",
    "build_manifest",
    "find_placeholders",
    "inspect_template",
    "list_templates",
    "placeholder_for",
    "repo_root",
    "resolve_template",
    "synthesize",
    "template_dirs",
]
