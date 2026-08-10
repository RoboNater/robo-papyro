"""The OOXML layer: namespaces, package mechanics, and the template content type.

openpyxl covers cells, formulas (as text), styles, tables, charts, images, and
most of the package on its own. What it does not cover — and the only reason
this module exists — is retyping a workbook saved under a different extension
(spec section 5.3) and the raw part inspection ``fidelity.py`` needs (section
6). The generic package mechanics underneath (zip read/repack, content-type
reading, the compiled-XPath helper) live in :mod:`rp_core.ooxml` and are
shared with the other OOXML leaves; this module wraps them with
SpreadsheetML's namespace map, content-type strings, and this package's own
errors — the same division ``rp_docx.ooxml`` and ``rp_pptx.ooxml`` use.

**Unlike ``rp_docx``/``rp_pptx``'s ``.dotx``/``.potx`` handling, there is no
read-side retyping to do.** openpyxl opens a ``.xltx``/``.xltm`` natively and
sets ``wb.template = True`` on load — verified,
``dev-notes/phase-3-openpyxl-probe.md`` section 6. The only workaround needed
is on the *write* side: ``wb.template`` is sticky, so a template-derived
workbook saved under a ``.xlsx`` name is still typed as a template unless
something resets the flag. :func:`save` is that something, driven by the
**output extension**, exactly where ``rp-docx`` and ``rp-pptx`` do their own
retyping.

Nothing here prints, and nothing here imports typer.
"""

from __future__ import annotations

import zipfile
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook

from rp_core import ooxml as core_ooxml
from rp_core.errors import InputError
from rp_xlsx.errors import InvalidXlsxError, MissingFileError

#: Every namespace this package resolves, in one place (spec section 7).
NS: dict[str, str] = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "ct": "http://schemas.openxmlformats.org/package/2006/content-types",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

#: openpyxl's own ``SUPPORTED_FORMATS`` (verified, spec section 1).
SUPPORTED_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")
MACRO_SUFFIXES = (".xlsm", ".xltm")
TEMPLATE_SUFFIXES = (".xltx", ".xltm")

WORKBOOK_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
TEMPLATE_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml"
)
MACRO_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
MACRO_TEMPLATE_CONTENT_TYPE = "application/vnd.ms-excel.template.macroEnabled.main+xml"

WORKBOOK_PART = "xl/workbook.xml"
VBA_PROJECT_PART = "xl/vbaProject.bin"

#: ``xpath(element, expr)`` with :data:`NS` bound (see
#: :func:`rp_core.ooxml.compiled_xpath` for why this goes through
#: ``etree.XPath`` rather than ``element.xpath(...)``).
xpath = core_ooxml.compiled_xpath(NS)


def format_of(path: Path) -> Literal["xlsx", "xlsm", "xltx", "xltm"]:
    """The format this package reports for ``path`` — its extension.

    Call after :func:`check_readable` has confirmed it is one of the four
    supported ones; this function does not itself validate.
    """
    return path.suffix.lower().lstrip(".")  # type: ignore[return-value]


def check_readable(path: Path) -> Path:
    """Fail early, and with the right error, on a path this package cannot open.

    Extension is checked **before** any content check (spec section 9,
    verified): a valid ``.xlsx`` renamed to ``.xls`` is refused on its name
    alone, and the message must say so rather than imply the file is corrupt.
    Legacy ``.xls``/``.xlsb`` get a distinct message naming the LibreOffice
    conversion command, because they are a documented non-goal rather than a
    bug this package could fix.
    """
    path = Path(path)
    if not path.exists():
        raise MissingFileError(f"No such file: {path}")
    if not path.is_file():
        raise MissingFileError(f"Not a file: {path}")
    suffix = path.suffix.lower()
    if suffix in (".xls", ".xlsb"):
        raise InvalidXlsxError(
            f"{path.name} has a {suffix} extension. rp-xlsx does not support legacy "
            f"{suffix} workbooks (a documented non-goal — openpyxl refuses them on the "
            f"extension alone). Convert it first: soffice --headless --convert-to xlsx "
            f"{path.name}"
        )
    if suffix not in SUPPORTED_SUFFIXES:
        raise InvalidXlsxError(
            f"{path.name} has a {suffix or '(no)'} extension; rp-xlsx opens "
            f"{', '.join(SUPPORTED_SUFFIXES)} workbooks."
        )
    if not zipfile.is_zipfile(path):
        raise InvalidXlsxError(f"{path.name} is not a valid OOXML package (not a zip archive).")
    return path


def part_names(path: Path) -> list[str]:
    """Every part in the package, in archive order."""
    check_readable(path)
    return core_ooxml.part_names(path)


def read_part(path: Path, name: str) -> bytes | None:
    """One part's bytes, or ``None`` when the package does not contain it."""
    check_readable(path)
    return core_ooxml.read_part(path, name)


def parse_part(path: Path, name: str) -> Any | None:
    """One part parsed as XML, or ``None`` when it is absent."""
    check_readable(path)
    try:
        return core_ooxml.parse_part(path, name)
    except ValueError as exc:
        raise InvalidXlsxError(f"{path.name}: {exc}.") from exc


def content_type_of(path: Path) -> str | None:
    """The main-part content type this package's zip actually declares.

    Used by tests to assert :func:`save`'s retyping at the byte level,
    independent of what a subsequent ``load_workbook`` reports — the two are
    supposed to agree, and a test that checked only one of them would miss a
    regression in whichever it did not check.
    """
    check_readable(path)
    return core_ooxml.content_type_from(
        path,
        (
            WORKBOOK_CONTENT_TYPE,
            TEMPLATE_CONTENT_TYPE,
            MACRO_WORKBOOK_CONTENT_TYPE,
            MACRO_TEMPLATE_CONTENT_TYPE,
        ),
    )


@contextmanager
def opened(path: Path, *, data_only: bool = False) -> Iterator[Workbook]:
    """An openpyxl ``Workbook`` for any of the four supported formats.

    ``keep_vba`` is selected from the extension — ``.xlsm``/``.xltm`` always
    load with ``keep_vba=True``, because the default silently drops
    ``xl/vbaProject.bin`` (verified) and there is no reading of "edit this
    workbook" that means "delete its code" (spec section 5.3). ``data_only``
    is the caller's to choose: reporting both a formula and its cached value
    needs two loads (spec section 9), so this function does not decide that.

    **The ``yield`` is deliberately inside a bare ``try``/``finally``, not
    ``try``/``except``** — only the open itself is ours to explain; a caller's
    own exception must propagate unchanged rather than being reported as a
    corrupt file (the same rule ``rp_pptx.ooxml.opened`` documents).
    """
    path = check_readable(path)
    keep_vba = path.suffix.lower() in MACRO_SUFFIXES
    try:
        workbook = load_workbook(str(path), data_only=data_only, keep_vba=keep_vba)
    except (zipfile.BadZipFile, InvalidFileException) as exc:
        raise InvalidXlsxError(f"Cannot open {path.name}: {exc}") from exc
    try:
        yield workbook
    finally:
        # `Workbook.close()` only affects read-only/write-only modes and
        # leaves `vba_archive` open — openpyxl keeps that handle alive so a
        # later `save()` can copy the VBA parts back out of it. Left
        # unclosed, its `ZipFile.__del__` eventually fires during garbage
        # collection, sometimes after the underlying file is already gone,
        # producing a spurious "I/O operation on closed file" warning.
        if workbook.vba_archive is not None:
            workbook.vba_archive.close()
        workbook.close()


def save(workbook: Workbook, output: Path) -> Path:
    """Save, retyping to a template when ``output`` names one, and setting
    ``fullCalcOnLoad`` unconditionally.

    Three things every write path needs, done in the one place every write
    path already funnels through:

    - **The output suffix must be one this package supports.** openpyxl's own
      ``save`` will happily write any filename; catching an unsupported one
      here means every write path refuses the same way, rather than each
      caller having to remember to check.
    - **Template retyping, from the output extension.** ``wb.template`` is
      sticky (verified, probe note section 6) — a workbook loaded from
      ``.xltx`` stays typed as a template even when saved under a ``.xlsx``
      name unless something resets the flag. This is that something.
    - **The macro/non-macro family must match the output extension.**
      ``Workbook.mime_type`` derives the content type from ``template`` *and*
      ``vba_archive`` together, and openpyxl's own docstring says it plainly:
      "Excel requires the file extension to match but openpyxl does not
      enforce this." Left unchecked, a macro-enabled workbook saved as
      ``.xlsx`` keeps its macros — and its macro-enabled content type —
      under a non-macro name, and a plain workbook saved as ``.xlsm`` gets a
      macro-enabled name over macro-free content. Neither is a case this
      package has a conversion story for, so both are refused rather than
      silently mislabeled.
    - **``fullCalcOnLoad = True``, on every save, unconditionally** (spec
      section 6.1). openpyxl's writer drops every formula's cached value; this
      is what makes Excel and LibreOffice recompute on open instead of a human
      seeing a workbook full of blanks. Not a flag — there is no reading of
      "save this workbook" that wants the alternative.
    """
    output = Path(output)
    suffix = output.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise InputError(
            f"{output.name} has a {suffix or '(no)'} extension; rp-xlsx writes "
            f"{', '.join(SUPPORTED_SUFFIXES)} workbooks."
        )
    has_macros = workbook.vba_archive is not None
    wants_macros = suffix in MACRO_SUFFIXES
    if has_macros and not wants_macros:
        raise InputError(
            f"{output.name} would drop this workbook's macros: openpyxl does not enforce "
            "that a file's extension matches its content, so saving a macro-enabled "
            f"workbook under a {suffix} name would silently keep the macro-enabled content "
            "type under a non-macro name. rp-xlsx refuses rather than mislabel it; strip "
            "the macros deliberately first if that is what you want."
        )
    if wants_macros and not has_macros:
        raise InputError(
            f"{output.name} is macro-enabled but the source workbook carries no macros; "
            "rp-xlsx will not write a macro-free file under a macro-enabled extension."
        )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.template = suffix in TEMPLATE_SUFFIXES
    workbook.calculation.fullCalcOnLoad = True
    workbook.save(str(output))
    return output


def populated_cells(ws: Any) -> list[Any]:
    """Every cell in ``ws`` that actually exists — holds a value, a formula,
    a comment, or a style — in row-major order.

    **Never scan a sheet with ``ws.iter_rows()`` and no bounds.** With no
    arguments it walks openpyxl's *declared* rectangle, ``1..max_row`` by
    ``1..max_col`` (``worksheet.py``'s own ``_cells_by_row``), constructing a
    ``Cell`` object for every position in between — section 9's
    phantom-dimension problem turned into a performance one, since a single
    stray formatted cell far from the real data (the ``phantom_dimension``
    fixture's own shape) makes a full-sheet scan cost ``max_row * max_col``
    cell constructions instead of the handful of cells the sheet actually
    has.

    ``Worksheet._cells`` holds exactly the cells that ever appeared in the
    sheet XML or were written to since (verified against openpyxl 3.1.5: a
    sheet with one value at A1 and only a fill at E1000 has exactly two
    entries, keyed ``(row, column)``), so reading it directly is both correct
    and cheap — a private attribute, not a public API, but the same one
    every performance-sensitive openpyxl consumer relies on for this reason.
    Sorted rather than trusted to already be in row-major order, because
    insertion order reflects how the sheet was populated, not its layout.
    """
    return [ws._cells[key] for key in sorted(ws._cells)]


def has_any_formula(workbook: Workbook) -> bool:
    """Whether any cell in ``workbook`` is a formula.

    Shared by every write path that needs to report
    ``recalculation_required``/``WriteResult.recalculation_required`` on its
    result (spec section 6.1): ``ooxml.save`` drops every formula's cached
    value on *every* save regardless of what a particular edit touched, so a
    caller learns whether that happened from the source's own formula count,
    not from what it wrote. Iterates the already-opened workbook rather than
    ``fidelity.has_cached_values`` — that answers a different question (does a
    *cached* value already exist on disk), and this one only needs to know
    whether a formula exists at all.
    """
    for ws in workbook.worksheets:
        for cell in populated_cells(ws):
            if cell.data_type == "f":
                return True
    return False


#: Every header/footer text slot openpyxl models on a worksheet — odd/even/
#: first, each with left/center/right. Shared by ``xlsx/write.py``'s
#: ``replace_text`` and ``xlsx/template.py``'s ``fill_template`` (spec
#: sections 4 and 8), which are the only two places this package touches
#: header/footer text.
HEADER_FOOTER_ATTRS = (
    "oddHeader",
    "evenHeader",
    "firstHeader",
    "oddFooter",
    "evenFooter",
    "firstFooter",
)
_HEADER_FOOTER_POSITIONS = ("left", "center", "right")


def header_footer_fields(ws: Any) -> list[tuple[str, Any]]:
    """Every populated header/footer text slot on ``ws``, as ``(label, part)``.

    ``part.text`` is the slot's current text (``get``/``set`` both work
    through it). The label is coarse (``"oddHeader"``, not
    ``"oddHeader:center"``) because that is what
    :class:`~rp_xlsx.models.ReplaceResult`'s ``locations`` reports.
    """
    fields: list[tuple[str, Any]] = []
    for attr in HEADER_FOOTER_ATTRS:
        container = getattr(ws, attr, None)
        if container is None:
            continue
        for position in _HEADER_FOOTER_POSITIONS:
            part = getattr(container, position, None)
            if part is not None:
                fields.append((attr, part))
    return fields


def require_output(output: Path | None) -> Path:
    """The resolved output path for an edit — never optional at this layer.

    Raises when ``output`` is ``None``: this package never overwrites an
    input file implicitly. The CLI turns ``--in-place`` into
    ``output=path`` explicitly, exactly as ``rp_pptx.ooxml.copy_for_edit``
    does for its own package.
    """
    if output is None:
        raise InputError(
            "An output path is required — this package never overwrites implicitly. "
            "Pass output=... (the CLI spells this -o OUT or --in-place)."
        )
    return Path(output)


__all__ = [
    "HEADER_FOOTER_ATTRS",
    "MACRO_SUFFIXES",
    "MACRO_TEMPLATE_CONTENT_TYPE",
    "MACRO_WORKBOOK_CONTENT_TYPE",
    "NS",
    "SUPPORTED_SUFFIXES",
    "TEMPLATE_CONTENT_TYPE",
    "TEMPLATE_SUFFIXES",
    "VBA_PROJECT_PART",
    "WORKBOOK_CONTENT_TYPE",
    "WORKBOOK_PART",
    "check_readable",
    "content_type_of",
    "format_of",
    "has_any_formula",
    "header_footer_fields",
    "opened",
    "parse_part",
    "part_names",
    "populated_cells",
    "read_part",
    "require_output",
    "save",
    "xpath",
]
