"""Writing workbooks: create, set_cells, append_rows, replace_text, set_properties.

**Every function that opens an existing workbook calls section 6's guard
first, before the workbook is opened.** A refusal costs nothing and a write
can never be left half-done. ``fullCalcOnLoad`` is set on every save via
``ooxml.save`` — never a flag, never optional — and
``WriteResult.recalculation_required``/``ReplaceResult.recalculation_required``
are computed from whether the *source* carried any formula, not from what a
particular edit touched, so a caller learns it from the result rather than
from documentation.

**A value beginning with `=` is written as a formula, always** — openpyxl's
own type inference does this. **A `'`-prefixed value is forced to text**,
mirroring Excel's own escape convention: the leading `'` is stripped and the
remainder (even if it still starts with `=`) is written as literal text,
never re-interpreted as a formula.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from rp_core.errors import InputError
from rp_xlsx import fidelity, ooxml, refs, templates
from rp_xlsx.models import CellValue, CoreProperties, ReplaceResult, SheetSpec, WriteResult

_BOLD = Font(bold=True)

#: Widths are derived from content length, capped so one long cell cannot
#: blow a sheet out to an unusable column (spec section 9).
_MIN_COLUMN_WIDTH = 6.0
_MAX_COLUMN_WIDTH = 60.0
_COLUMN_PADDING = 2.0


def _set_cell_value(cell: Any, value: CellValue) -> None:
    """Write ``value`` into ``cell``, honouring the `'`-escape (spec section 4).

    openpyxl does **not** strip a leading apostrophe itself (verified) — left
    alone, ``"'=foo"`` would be stored, apostrophe and all, as the literal
    text ``"'=foo"`` rather than Excel's ``"=foo"``. Replicating Excel's
    convention means doing the stripping here and then forcing the data type,
    because reassigning a string that still starts with ``=`` would otherwise
    make openpyxl re-detect it as a formula.
    """
    if isinstance(value, str) and value.startswith("'"):
        cell.value = value[1:]
        cell.data_type = "s"
    else:
        cell.value = value


def _sheet_by_name(wb: Any, name: str) -> Any:
    if name not in wb.sheetnames:
        available = ", ".join(repr(n) for n in wb.sheetnames)
        raise InputError(f"No sheet named {name!r}. Available sheets: {available}")
    return wb[name]


def _next_empty_row(ws: Any) -> int:
    """The first row below every row holding a value or a formula.

    Never ``ws.max_row + 1``: that inherits section 9's phantom-dimension
    lie (a format-only cell far below the data inflates ``max_row``), which
    would make ``append_rows`` open a gap of blank rows instead of appending.
    Scans ``ooxml.populated_cells`` rather than ``ws.iter_rows()`` for the
    same reason ``read._used_bounds`` does: the declared rectangle is exactly
    what a phantom dimension inflates, so walking it would trade the wrong
    answer for a slow one instead of fixing it.
    """
    last = 0
    for cell in ooxml.populated_cells(ws):
        if cell.value is not None:
            last = max(last, cell.row)
    return last + 1


def _autosized_widths(header: list[str] | None, rows: list[list[CellValue]]) -> dict[str, float]:
    widest: dict[int, int] = {}
    for row in ([header] if header else []) + rows:
        for col, value in enumerate(row, start=1):
            length = 0 if value is None else len(str(value))
            widest[col] = max(widest.get(col, 0), length)
    return {
        refs.column_letters(col): min(max(length, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
        for col, length in ((c, w + _COLUMN_PADDING) for c, w in widest.items())
    }


def _populate_sheet(ws: Any, spec: SheetSpec, header_style: bool) -> int:
    cells_written = 0
    row_index = 1
    if spec.header:
        for col, value in enumerate(spec.header, start=1):
            _set_cell_value(ws.cell(row=row_index, column=col), value)
            cells_written += 1
        if header_style:
            for col in range(1, len(spec.header) + 1):
                ws.cell(row=row_index, column=col).font = _BOLD
            ws.freeze_panes = "A2"
        elif spec.freeze_header:
            ws.freeze_panes = "A2"
        row_index += 1
    for row in spec.rows:
        for col, value in enumerate(row, start=1):
            _set_cell_value(ws.cell(row=row_index, column=col), value)
            cells_written += 1
        row_index += 1
    widths = spec.column_widths or _autosized_widths(spec.header, spec.rows)
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    return cells_written


def _populate_workbook(
    wb: Any, sheets: list[SheetSpec] | None, header_style: bool, *, replacing_default: bool
) -> int:
    if not sheets:
        return 0
    seen: set[str] = set()
    seen_folded: set[str] = set()
    for spec in sheets:
        folded = spec.name.casefold()
        if folded in seen_folded:
            raise InputError(
                f"Duplicate sheet name {spec.name!r} in the same create() call "
                "(sheet names collide case-insensitively)."
            )
        seen.add(spec.name)
        seen_folded.add(folded)
    original_default = wb.sheetnames[0] if replacing_default else None
    cells_written = 0
    for spec in sheets:
        others = [name for name in wb.sheetnames if name != spec.name]
        refs.validate_sheet_name(spec.name, others)
        if spec.name in wb.sheetnames:
            del wb[spec.name]
        ws = wb.create_sheet(spec.name)
        cells_written += _populate_sheet(ws, spec, header_style)
    if original_default and original_default not in seen and original_default in wb.sheetnames:
        del wb[original_default]
    return cells_written


def create(
    output: Path,
    *,
    sheets: list[SheetSpec] | None = None,
    template: str | Path | None = None,
    header_style: bool = True,
) -> Path:
    """Build a new workbook, optionally starting from a template's shell.

    ``template=None`` starts from ``openpyxl.Workbook()`` — there is no
    bundled default to fall back to (spec section 4's considered divergence
    from ``rp-docx``/``rp-pptx``). A resolved template's own defined names,
    styles, and other sheets are kept; each ``SheetSpec`` becomes a freshly
    written sheet, replacing any template sheet of the same name outright
    rather than attempting a partial cell-by-cell overlay. Filling a
    template's own placeholder cells in place is ``fill_template``'s job
    (spec section 8), not this function's.
    """
    output = Path(output)
    resolved = templates.resolve_template(template)
    if resolved is not None:
        with ooxml.opened(resolved) as wb:
            _populate_workbook(wb, sheets, header_style, replacing_default=False)
            return ooxml.save(wb, output)

    # A blank ``Workbook()`` never carries macros, so a macro-enabled output
    # suffix here is refused by ``ooxml.save``'s own macro/non-macro check —
    # the same check that also covers a template-backed create against a
    # mismatched suffix, which a check only in this branch would have missed.
    wb = Workbook()
    _populate_workbook(wb, sheets, header_style, replacing_default=True)
    return ooxml.save(wb, output)


def set_cells(
    path: Path,
    updates: dict[str, dict[str, CellValue]],
    *,
    output: Path | None = None,
    allow_lossy: bool = False,
) -> WriteResult:
    """Set specific cells: ``{"Sheet1": {"B2": 5, "C3": "=B2*2"}}``."""
    report = fidelity.guard(path, allow_lossy=allow_lossy)
    target = ooxml.require_output(output)
    with ooxml.opened(path) as wb:
        recalculation_required = ooxml.has_any_formula(wb)
        cells_written = 0
        for sheet_name, cell_updates in updates.items():
            ws = _sheet_by_name(wb, sheet_name)
            for ref, value in cell_updates.items():
                position = refs.parse_cell_ref(ref)
                _set_cell_value(ws.cell(row=position.row, column=position.column), value)
                cells_written += 1
        ooxml.save(wb, target)
    return WriteResult(
        output=target,
        cells_written=cells_written,
        recalculation_required=recalculation_required,
        dropped=report.at_risk,
    )


def append_rows(
    path: Path,
    sheet: str,
    rows: list[list[CellValue]],
    *,
    output: Path | None = None,
    allow_lossy: bool = False,
) -> WriteResult:
    """Append rows after the sheet's last used row (never after its
    possibly-phantom declared dimension — spec section 9)."""
    report = fidelity.guard(path, allow_lossy=allow_lossy)
    target = ooxml.require_output(output)
    with ooxml.opened(path) as wb:
        recalculation_required = ooxml.has_any_formula(wb)
        ws = _sheet_by_name(wb, sheet)
        start_row = _next_empty_row(ws)
        cells_written = 0
        for offset, row in enumerate(rows):
            for col, value in enumerate(row, start=1):
                _set_cell_value(ws.cell(row=start_row + offset, column=col), value)
                cells_written += 1
        ooxml.save(wb, target)
    return WriteResult(
        output=target,
        cells_written=cells_written,
        recalculation_required=recalculation_required,
        dropped=report.at_risk,
    )


def _replace_in_text(
    text: str, replacements: dict[str, str], order: list[str], match_case: bool
) -> tuple[str, dict[str, int]]:
    """``text`` with every key in ``order`` substituted, longest-first so
    overlapping keys resolve deterministically regardless of dict order."""
    hits: dict[str, int] = {}
    result = text
    for key in order:
        if match_case:
            count = result.count(key)
            if count:
                hits[key] = count
                result = result.replace(key, replacements[key])
        else:
            pattern = re.compile(re.escape(key), re.IGNORECASE)
            count = len(pattern.findall(result))
            if count:
                hits[key] = count
                replacement = replacements[key]
                result = pattern.sub(lambda _m, r=replacement: r, result)
    return result, hits


def replace_text(
    path: Path,
    replacements: dict[str, str],
    *,
    output: Path | None = None,
    sheets: str = "all",
    match_case: bool = True,
    include_formulas: bool = False,
    allow_lossy: bool = False,
) -> ReplaceResult:
    """Replace text in cell values (and header/footer text) on the selected
    sheets. Does not touch formulas unless ``include_formulas=True`` (spec
    section 4): a replacement landing inside ``=SUM(Revenue!A1:A9)`` would
    otherwise produce a formula that is broken or silently pointing
    somewhere else.
    """
    report = fidelity.guard(path, allow_lossy=allow_lossy)
    target = ooxml.require_output(output)
    counts = dict.fromkeys(replacements, 0)
    locations: list[str] = []
    order = sorted(replacements, key=len, reverse=True)
    with ooxml.opened(path) as wb:
        recalculation_required = ooxml.has_any_formula(wb)
        positions = refs.resolve_sheet_selection(wb.sheetnames, sheets=sheets)
        for position in positions:
            ws = wb.worksheets[position - 1]
            for cell in ooxml.populated_cells(ws):
                if cell.data_type == "f" and not include_formulas:
                    continue
                if not isinstance(cell.value, str):
                    continue
                new_text, hits = _replace_in_text(cell.value, replacements, order, match_case)
                if not hits:
                    continue
                cell.value = new_text
                for key, n in hits.items():
                    counts[key] += n
                locations.append(f"{ws.title}!{cell.coordinate}")
            header_hit = False
            for _label, part in ooxml.header_footer_fields(ws):
                if not part.text:
                    continue
                new_text, hits = _replace_in_text(part.text, replacements, order, match_case)
                if not hits:
                    continue
                part.text = new_text
                header_hit = True
                for key, n in hits.items():
                    counts[key] += n
            if header_hit:
                locations.append(f"header:{ws.title}")
        ooxml.save(wb, target)
    return ReplaceResult(
        output=target,
        replacements=counts,
        locations=locations,
        recalculation_required=recalculation_required,
        dropped=report.at_risk,
    )


def set_properties(
    path: Path,
    props: CoreProperties,
    *,
    output: Path | None = None,
    allow_lossy: bool = False,
) -> WriteResult:
    """Update core properties. ``None`` fields mean leave alone, not clear —
    openpyxl's ``created``/``modified`` are non-nullable on save, so blindly
    assigning every field would crash on the common case of a caller setting
    only ``title``.

    Returns a :class:`~rp_xlsx.models.WriteResult`, exactly like every other
    edit of an existing workbook (spec section 6's contract applies here too:
    a workbook is opened and re-saved, so its formulas lose their cached
    values and any at-risk part named by ``allow_lossy`` is just as real a
    loss as it is for ``set_cells``). ``cells_written`` is always 0 — this
    function touches workbook-level properties, never a cell.
    """
    report = fidelity.guard(path, allow_lossy=allow_lossy)
    target = ooxml.require_output(output)
    field_map = {
        "title": "title",
        "author": "creator",
        "last_modified_by": "lastModifiedBy",
        "created": "created",
        "modified": "modified",
        "category": "category",
        "keywords": "keywords",
    }
    with ooxml.opened(path) as wb:
        recalculation_required = ooxml.has_any_formula(wb)
        for field, attr in field_map.items():
            value = getattr(props, field)
            if value is not None:
                setattr(wb.properties, attr, value)
        if props.revision is not None:
            wb.properties.revision = str(props.revision)
        ooxml.save(wb, target)
    return WriteResult(
        output=target,
        cells_written=0,
        recalculation_required=recalculation_required,
        dropped=report.at_risk,
    )


__all__ = [
    "append_rows",
    "create",
    "replace_text",
    "set_cells",
    "set_properties",
]
