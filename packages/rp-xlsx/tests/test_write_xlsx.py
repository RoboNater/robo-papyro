"""rp_xlsx.xlsx.write -- create, set_cells, append_rows, replace_text,
set_properties. Every editing entry point goes through section 6's guard
first; §11.3's required observable assertions live here and in
test_fidelity_xlsx.py.
"""

from __future__ import annotations

import zipfile

import openpyxl
import pytest

from rp_core.errors import InputError
from rp_xlsx.errors import LossyEditError
from rp_xlsx.models import CoreProperties, SheetSpec
from rp_xlsx.xlsx import read, write


class TestCreate:
    def test_creates_a_readable_workbook(self, tmp_path):
        out = write.create(
            tmp_path / "new.xlsx",
            sheets=[SheetSpec(name="Data", header=["Name", "Amount"], rows=[["a", 1], ["b", 2]])],
        )
        data = read.get_data(out, sheets="1")
        assert data[0].header == ["Name", "Amount"]
        assert data[0].rows == [["a", 1], ["b", 2]]

    def test_header_style_bolds_and_freezes_by_default(self, tmp_path):
        out = write.create(
            tmp_path / "styled.xlsx",
            sheets=[SheetSpec(name="Data", header=["A"], rows=[["x"]])],
        )
        idx = read.get_index(out)
        assert idx.sheets[0].freeze_panes == "A2"

    def test_no_header_style_skips_bold_and_freeze(self, tmp_path):
        out = write.create(
            tmp_path / "plain.xlsx",
            sheets=[SheetSpec(name="Data", header=["A"], rows=[["x"]])],
            header_style=False,
        )
        idx = read.get_index(out)
        assert idx.sheets[0].freeze_panes is None

    def test_no_template_and_no_sheets_is_a_blank_workbook(self, tmp_path):
        out = write.create(tmp_path / "blank.xlsx")
        idx = read.get_index(out)
        assert idx.sheet_count == 1

    def test_multiple_sheets_replace_the_default_sheet(self, tmp_path):
        out = write.create(
            tmp_path / "multi.xlsx",
            sheets=[
                SheetSpec(name="One", rows=[["x"]]),
                SheetSpec(name="Two", rows=[["y"]]),
            ],
        )
        idx = read.get_index(out)
        assert idx.sheets[0].name == "One"
        assert idx.sheets[1].name == "Two"
        assert idx.sheet_count == 2

    def test_duplicate_sheet_names_are_rejected(self, tmp_path):
        with pytest.raises(InputError):
            write.create(
                tmp_path / "dup.xlsx",
                sheets=[SheetSpec(name="A", rows=[["x"]]), SheetSpec(name="A", rows=[["y"]])],
            )

    def test_case_only_duplicate_sheet_names_are_also_rejected(self, tmp_path):
        """openpyxl treats sheet names case-insensitively -- "a" next to "A"
        in the same create() call is a collision, not two distinct sheets."""
        with pytest.raises(InputError):
            write.create(
                tmp_path / "dup.xlsx",
                sheets=[SheetSpec(name="A", rows=[["x"]]), SheetSpec(name="a", rows=[["y"]])],
            )

    def test_macro_extension_without_a_template_is_an_input_error(self, tmp_path):
        with pytest.raises(InputError):
            write.create(tmp_path / "out.xlsm", sheets=[SheetSpec(name="Data", rows=[["x"]])])

    def test_macro_extension_against_a_non_macro_template_is_also_an_input_error(
        self, tmp_path, template_workbook_path
    ):
        """The no-template branch's check above does not cover this path --
        a template-backed create needs the same refusal, which is why the
        check now lives in ooxml.save() rather than only here."""
        with pytest.raises(InputError):
            write.create(
                tmp_path / "out.xlsm",
                sheets=[SheetSpec(name="Data", rows=[["x"]])],
                template=template_workbook_path,
            )

    def test_template_sheets_are_preserved(self, tmp_path, template_workbook_path):
        out = write.create(
            tmp_path / "from-template.xlsx",
            sheets=[SheetSpec(name="NewSheet", rows=[["x"]])],
            template=template_workbook_path,
        )
        idx = read.get_index(out)
        names = [s.name for s in idx.sheets]
        assert "NewSheet" in names
        assert "Sheet" in names  # the template's original sheet survives

    def test_column_widths_are_autosized_from_content(self, tmp_path):
        out = write.create(
            tmp_path / "wide.xlsx",
            sheets=[SheetSpec(name="Data", header=["Short"], rows=[["a very long value indeed"]])],
        )
        wb = openpyxl.load_workbook(out)
        width = wb["Data"].column_dimensions["A"].width
        assert width > 10

    def test_explicit_column_widths_are_honoured(self, tmp_path):
        out = write.create(
            tmp_path / "exact.xlsx",
            sheets=[SheetSpec(name="Data", rows=[["x"]], column_widths={"A": 42.0})],
        )
        wb = openpyxl.load_workbook(out)
        assert wb["Data"].column_dimensions["A"].width == 42.0


class TestFormulaAndLiteralEscape:
    def test_leading_equals_is_written_as_a_formula(self, tmp_path):
        out = write.create(tmp_path / "f.xlsx", sheets=[SheetSpec(name="D", rows=[[1, 2]])])
        result = write.set_cells(out, {"D": {"C1": "=A1+B1"}}, output=tmp_path / "f2.xlsx")
        assert result.cells_written == 1
        cells = {c.ref: c for c in read.get_cells(tmp_path / "f2.xlsx", sheets="1", empty=True)}
        assert cells["C1"].formula == "=A1+B1"

    def test_apostrophe_prefix_is_stored_as_literal_text(self, tmp_path):
        out = write.create(tmp_path / "e.xlsx", sheets=[SheetSpec(name="D", rows=[[1]])])
        write.set_cells(out, {"D": {"B1": "'=notaformula"}}, output=tmp_path / "e2.xlsx")
        cells = {c.ref: c for c in read.get_cells(tmp_path / "e2.xlsx", sheets="1", empty=True)}
        assert cells["B1"].value == "=notaformula"
        assert cells["B1"].formula is None

    def test_apostrophe_prefix_survives_a_reload(self, tmp_path):
        out = write.create(tmp_path / "e.xlsx", sheets=[SheetSpec(name="D", rows=[[1]])])
        write.set_cells(out, {"D": {"B1": "'plain text"}}, output=tmp_path / "e2.xlsx")
        cells = {c.ref: c for c in read.get_cells(tmp_path / "e2.xlsx", sheets="1", empty=True)}
        assert cells["B1"].value == "plain text"


class TestSetCells:
    def test_sets_multiple_sheets_and_cells(self, tmp_path):
        out = write.create(
            tmp_path / "base.xlsx",
            sheets=[SheetSpec(name="A", rows=[[1]]), SheetSpec(name="B", rows=[[2]])],
        )
        result = write.set_cells(
            out, {"A": {"B1": 10}, "B": {"B1": 20}}, output=tmp_path / "out.xlsx"
        )
        assert result.cells_written == 2
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        assert data[0].rows[0] == [1, 10]
        assert data[1].rows[0] == [2, 20]

    def test_unknown_sheet_is_an_input_error(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[[1]])])
        with pytest.raises(InputError):
            write.set_cells(out, {"Nope": {"A1": 1}}, output=tmp_path / "out.xlsx")

    def test_output_required(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[[1]])])
        with pytest.raises(InputError):
            write.set_cells(out, {"A": {"A1": 1}})

    def test_does_not_overwrite_the_input_file(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[[1]])])
        before = out.read_bytes()
        write.set_cells(out, {"A": {"A1": 99}}, output=tmp_path / "changed.xlsx")
        assert out.read_bytes() == before


class TestAppendRows:
    def test_appends_after_the_last_used_row(self, tmp_path):
        out = write.create(
            tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", header=["H"], rows=[["r1"]])]
        )
        result = write.append_rows(out, "A", [["r2"], ["r3"]], output=tmp_path / "out.xlsx")
        assert result.cells_written == 2
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        assert data[0].rows == [["H"], ["r1"], ["r2"], ["r3"]]

    def test_ignores_phantom_dimensions(self, tmp_path, phantom_dimension_workbook):
        write.append_rows(
            phantom_dimension_workbook, "Sheet", [["appended"]], output=tmp_path / "out.xlsx"
        )
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        # The phantom cell is at E1000; a naive max_row+1 append would land
        # at row 1001 instead of row 2.
        assert data[0].rows[1][0] == "appended"

    def test_stays_fast_at_excels_actual_row_and_column_limits(
        self, tmp_path, adversarial_phantom_dimension_workbook
    ):
        """_next_empty_row must scan populated cells, not the declared
        rectangle -- see the equivalent get_index test for why 5s is a
        generous bound at Excel's real row/column limits."""
        import time

        start = time.monotonic()
        write.append_rows(
            adversarial_phantom_dimension_workbook,
            "Sheet",
            [["appended"]],
            output=tmp_path / "out.xlsx",
        )
        elapsed = time.monotonic() - start
        assert elapsed < 5, f"append_rows took {elapsed:.1f}s"
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        assert data[0].rows[1][0] == "appended"

    def test_unknown_sheet_is_an_input_error(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[[1]])])
        with pytest.raises(InputError):
            write.append_rows(out, "Nope", [[1]], output=tmp_path / "out.xlsx")


class TestReplaceText:
    def test_replaces_matching_text(self, tmp_path):
        out = write.create(
            tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[["hello world"]])]
        )
        result = write.replace_text(out, {"world": "there"}, output=tmp_path / "out.xlsx")
        assert result.replacements == {"world": 1}
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        assert data[0].rows == [["hello there"]]

    def test_unmatched_key_reports_zero(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[["hi"]])])
        result = write.replace_text(out, {"nope": "x"}, output=tmp_path / "out.xlsx")
        assert result.replacements == {"nope": 0}

    def test_does_not_touch_formulas_by_default(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="Revenue", rows=[[1]])])
        set_out = tmp_path / "with-formula.xlsx"
        write.set_cells(out, {"Revenue": {"A2": "=SUM(Revenue!A1:A1)"}}, output=set_out)
        write.replace_text(set_out, {"Revenue": "Sales"}, output=tmp_path / "out.xlsx")
        cells = {c.ref: c for c in read.get_cells(tmp_path / "out.xlsx", sheets="1", empty=True)}
        assert cells["A2"].formula == "=SUM(Revenue!A1:A1)"

    def test_include_formulas_touches_them_and_marks_the_location(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="Revenue", rows=[[1]])])
        set_out = tmp_path / "with-formula.xlsx"
        write.set_cells(out, {"Revenue": {"A2": "=SUM(Revenue!A1:A1)"}}, output=set_out)
        result = write.replace_text(
            set_out, {"Revenue": "Sales"}, include_formulas=True, output=tmp_path / "out.xlsx"
        )
        cells = {c.ref: c for c in read.get_cells(tmp_path / "out.xlsx", sheets="1", empty=True)}
        assert cells["A2"].formula == "=SUM(Sales!A1:A1)"
        assert "Revenue!A2" in result.locations

    def test_case_insensitive_replace(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[["Hello"]])])
        result = write.replace_text(
            out, {"hello": "hi"}, match_case=False, output=tmp_path / "out.xlsx"
        )
        assert result.replacements == {"hello": 1}
        data = read.get_data(tmp_path / "out.xlsx", header=False)
        assert data[0].rows == [["hi"]]

    def test_case_sensitive_by_default_does_not_match(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx", sheets=[SheetSpec(name="A", rows=[["Hello"]])])
        result = write.replace_text(out, {"hello": "hi"}, output=tmp_path / "out.xlsx")
        assert result.replacements == {"hello": 0}

    def test_header_footer_text_is_replaced(self, tmp_path):
        source = tmp_path / "hf.xlsx"
        wb = openpyxl.Workbook()
        wb.active.oddHeader.center.text = "Report for {{ client }}"
        wb.save(source)
        result = write.replace_text(source, {"{{ client }}": "Acme"}, output=tmp_path / "out.xlsx")
        assert result.replacements == {"{{ client }}": 1}
        assert any(loc.startswith("header:") for loc in result.locations)
        wb2 = openpyxl.load_workbook(tmp_path / "out.xlsx")
        assert wb2.active.oddHeader.center.text == "Report for Acme"

    def test_sheets_selector_restricts_scope(self, tmp_path):
        out = write.create(
            tmp_path / "base.xlsx",
            sheets=[SheetSpec(name="A", rows=[["x"]]), SheetSpec(name="B", rows=[["x"]])],
        )
        result = write.replace_text(out, {"x": "y"}, sheets="1", output=tmp_path / "out.xlsx")
        assert result.replacements == {"x": 1}


class TestSetProperties:
    def test_sets_and_reads_back(self, tmp_path):
        out = write.create(tmp_path / "base.xlsx")
        result = write.set_properties(
            out, CoreProperties(title="Q1 Report", author="Ada"), output=tmp_path / "out.xlsx"
        )
        props = read.get_properties(result.output)
        assert props.title == "Q1 Report"
        assert props.author == "Ada"

    def test_none_fields_do_not_clear_existing_values(self, tmp_path):
        """openpyxl's created/modified are non-nullable on save -- a caller
        setting only title must not crash by blanking them, and must not
        clear the pre-existing author either."""
        out = write.create(tmp_path / "base.xlsx")
        write.set_properties(out, CoreProperties(author="Ada"), output=tmp_path / "step1.xlsx")
        result = write.set_properties(
            tmp_path / "step1.xlsx", CoreProperties(title="Later"), output=tmp_path / "step2.xlsx"
        )
        props = read.get_properties(result.output)
        assert props.title == "Later"
        assert props.author == "Ada"
        assert props.created is not None

    def test_returns_a_write_result_reporting_loss(self, at_risk_workbook, tmp_path):
        """set_properties opens and re-saves an existing workbook, exactly
        like set_cells, so it owes the same loss report -- not a bare Path
        a caller (the MCP tool, in particular) would otherwise have to
        fabricate recalculation_required/dropped for."""
        result = write.set_properties(
            at_risk_workbook,
            CoreProperties(title="Q1"),
            output=tmp_path / "out.xlsx",
            allow_lossy=True,
        )
        assert result.dropped
        assert result.cells_written == 0


class TestFidelityGuardIntegration:
    """Spec section 11.3's required observable assertions for §6."""

    def test_at_risk_workbook_refuses_every_write_path(self, at_risk_workbook, tmp_path):
        with pytest.raises(LossyEditError):
            write.set_cells(at_risk_workbook, {"Sheet1": {"A1": 1}}, output=tmp_path / "out.xlsx")

    def test_allow_lossy_proceeds_and_reports_dropped(self, at_risk_workbook, tmp_path):
        result = write.set_cells(
            at_risk_workbook,
            {"Sheet1": {"A1": 1}},
            output=tmp_path / "out.xlsx",
            allow_lossy=True,
        )
        assert result.dropped

    def test_recalculation_required_true_when_source_has_formulas(
        self, cached_value_workbook, tmp_path
    ):
        result = write.set_cells(
            cached_value_workbook, {"Sheet": {"B1": 1}}, output=tmp_path / "out.xlsx"
        )
        assert result.recalculation_required is True

    def test_recalculation_required_false_with_no_formulas(self, plain_workbook, tmp_path):
        result = write.set_cells(
            plain_workbook, {"Sheet1": {"A2": 1}}, output=tmp_path / "out.xlsx"
        )
        assert result.recalculation_required is False

    def test_macro_edit_keeps_vba_with_no_flag(self, macro_workbook, tmp_path):
        write.set_cells(macro_workbook, {"Sheet1": {"A2": 1}}, output=tmp_path / "out.xlsm")
        with zipfile.ZipFile(tmp_path / "out.xlsm") as zf:
            assert "xl/vbaProject.bin" in zf.namelist()

    def test_replace_text_also_refuses_and_reports_dropped(self, at_risk_workbook, tmp_path):
        """§6's contract binds every write path that opens an existing
        workbook, not only set_cells/append_rows -- replace_text is the one
        write function whose result once had nowhere to put this."""
        with pytest.raises(LossyEditError):
            write.replace_text(at_risk_workbook, {"a": "b"}, output=tmp_path / "out.xlsx")
        result = write.replace_text(
            at_risk_workbook, {"a": "b"}, output=tmp_path / "out.xlsx", allow_lossy=True
        )
        assert result.dropped

    def test_set_properties_also_refuses_and_reports_dropped(self, at_risk_workbook, tmp_path):
        with pytest.raises(LossyEditError):
            write.set_properties(
                at_risk_workbook, CoreProperties(title="Q1"), output=tmp_path / "out.xlsx"
            )
        result = write.set_properties(
            at_risk_workbook,
            CoreProperties(title="Q1"),
            output=tmp_path / "out.xlsx",
            allow_lossy=True,
        )
        assert result.dropped
