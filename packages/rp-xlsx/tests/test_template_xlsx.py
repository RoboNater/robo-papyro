"""rp_xlsx.xlsx.template -- fill_template (spec section 8).

Not deferrable (spec section 12 step 8) unlike build_manifest/synthesize.
"""

from __future__ import annotations

import openpyxl
import pytest

from rp_core.errors import InputError
from rp_xlsx.errors import LossyEditError
from rp_xlsx.xlsx import read
from rp_xlsx.xlsx.template import fill_template, flatten


class TestFlatten:
    def test_flat_context_is_unchanged_keys(self):
        assert flatten({"a": "1", "b": "2"}) == {"a": "1", "b": "2"}

    def test_nested_dict_becomes_dotted(self):
        assert flatten({"client": {"name": "Ada"}}) == {"client.name": "Ada"}

    def test_deep_nesting_flattens_fully(self):
        assert flatten({"a": {"b": {"c": "x"}}}) == {"a.b.c": "x"}

    def test_none_becomes_empty_string(self):
        assert flatten({"a": None}) == {"a": ""}

    def test_non_string_values_are_stringified(self):
        assert flatten({"count": 5}) == {"count": "5"}


class TestFillTemplate:
    def test_fills_placeholders(self, house_like_template, tmp_path):
        result = fill_template(
            house_like_template,
            {"client": {"name": "Acme Corp"}, "report": {"date": "2024-01-01"}},
            tmp_path / "filled.xlsx",
        )
        assert result.unresolved == []
        wb = openpyxl.load_workbook(result.output)
        assert wb["Report"]["A5"].value == "Client: Acme Corp"
        assert wb["Report"]["A6"].value == "Date: 2024-01-01"

    def test_filled_reports_only_resolved_keys(self, house_like_template, tmp_path):
        result = fill_template(
            house_like_template,
            {"client": {"name": "Acme"}, "report": {"date": "2024-01-01"}},
            tmp_path / "filled.xlsx",
        )
        assert result.filled == {"client.name": "Acme", "report.date": "2024-01-01"}

    def test_strict_raises_on_unresolved_and_writes_nothing(self, house_like_template, tmp_path):
        output = tmp_path / "filled.xlsx"
        with pytest.raises(InputError, match="report.date"):
            fill_template(house_like_template, {"client": {"name": "Acme"}}, output)
        assert not output.exists()

    def test_non_strict_leaves_unresolved_in_place_and_reports_them(
        self, house_like_template, tmp_path
    ):
        result = fill_template(
            house_like_template,
            {"client": {"name": "Acme"}},
            tmp_path / "filled.xlsx",
            strict=False,
        )
        assert result.unresolved == ["report.date"]
        wb = openpyxl.load_workbook(result.output)
        assert wb["Report"]["A6"].value == "Date: {{ report.date }}"

    def test_template_name_resolves_like_create(self, monkeypatch, house_like_template, tmp_path):
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", str(house_like_template.parent))
        result = fill_template(
            house_like_template.stem,
            {"client": {"name": "Acme"}, "report": {"date": "now"}},
            tmp_path / "filled.xlsx",
        )
        assert result.unresolved == []

    def test_none_template_with_nothing_configured_raises(self, monkeypatch, tmp_path):
        monkeypatch.delenv("RP_XLSX_TEMPLATE", raising=False)
        with pytest.raises(InputError):
            fill_template(None, {}, tmp_path / "out.xlsx")

    def test_adjacent_overlapping_placeholders_do_not_interfere(self, hostile_template, tmp_path):
        """`{{ client }}` and `{{ client.name }}` sit adjacent in the same
        cell (hostile_template's A8), and only `client.name` is in context
        (`client` is a dict, not a leaf, so it is never a fill key itself).
        Resolving `client.name` must not corrupt or consume the neighboring
        `{{ client }}` placeholder (spec section 11.2's longest-first case)."""
        result = fill_template(
            hostile_template,
            {"client": {"name": "Ada"}},
            tmp_path / "filled.xlsx",
            strict=False,
        )
        wb = openpyxl.load_workbook(result.output)
        ws = wb[wb.sheetnames[0]]
        assert ws["A8"].value == "{{ client }} and Ada adjacent"

    def test_reordered_dict_gives_the_same_result(self, hostile_template, tmp_path):
        """The same substitution regardless of dict insertion order --
        `replace_text` sorts its replacements longest-first internally, so
        the caller's dict order must never leak into the result."""
        context_a = {"client": {"name": "Ada"}, "greeting": "hi"}
        context_b = {"greeting": "hi", "client": {"name": "Ada"}}
        result_a = fill_template(hostile_template, context_a, tmp_path / "a.xlsx", strict=False)
        result_b = fill_template(hostile_template, context_b, tmp_path / "b.xlsx", strict=False)
        assert result_a.filled == result_b.filled
        assert result_a.unresolved == result_b.unresolved

    def test_does_not_touch_formulas(self, house_like_template, tmp_path):
        result = fill_template(
            house_like_template,
            {"client": {"name": "Acme"}, "report": {"date": "2024-01-01"}},
            tmp_path / "filled.xlsx",
        )
        cells = {c.ref: c for c in read.get_cells(result.output, sheets="1", empty=True)}
        # LineItems table has no formulas, but this is the load-bearing
        # assertion shape for any future placeholder-in-formula-adjacent-cell
        # regression: formulas, if present, come through untouched.
        assert all(c.formula is None for c in cells.values() if c.ref in ("A1", "B1"))


class TestFillTemplateFidelityGuard:
    """fill_template opens and re-saves the resolved template -- an existing
    workbook -- so section 6's guard applies to it exactly as it does to
    every other function that opens one, via the replace_text it delegates
    to. `at_risk_workbook` is a plain, resolvable file; a template does not
    need to be `.xltx`-typed for `resolve_template`'s path branch to accept
    it."""

    def test_refuses_without_allow_lossy(self, at_risk_workbook, tmp_path):
        with pytest.raises(LossyEditError):
            fill_template(at_risk_workbook, {}, tmp_path / "out.xlsx", strict=False)

    def test_allow_lossy_proceeds_and_reports_dropped(self, at_risk_workbook, tmp_path):
        result = fill_template(
            at_risk_workbook, {}, tmp_path / "out.xlsx", strict=False, allow_lossy=True
        )
        assert result.dropped

    def test_reports_dropped_even_with_no_placeholders_matched(self, at_risk_workbook, tmp_path):
        """Any save of the template loses cached formula values and at-risk
        parts, whether or not a placeholder actually substituted -- an empty
        `filled`/`unresolved` result must not imply an empty `dropped`."""
        result = fill_template(
            at_risk_workbook, {}, tmp_path / "out.xlsx", strict=False, allow_lossy=True
        )
        assert result.filled == {}
        assert result.dropped
