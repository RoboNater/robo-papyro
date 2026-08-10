"""rp_xlsx.xlsx.sheets -- add, delete, rename, reorder.

Section 4's rules: the permutation check for reorder (rp-pptx's rule,
unchanged), the at-least-one-visible-sheet invariant (stricter than
openpyxl, which will happily write a workbook Excel refuses to open), and
sheet-name validation on the way in.
"""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from rp_core.errors import InputError
from rp_xlsx.errors import LossyEditError
from rp_xlsx.models import SheetSpec
from rp_xlsx.xlsx import read, sheets, write


@pytest.fixture
def three_sheet_workbook(tmp_path):
    return write.create(
        tmp_path / "three.xlsx",
        sheets=[
            SheetSpec(name="One", rows=[["1"]]),
            SheetSpec(name="Two", rows=[["2"]]),
            SheetSpec(name="Three", rows=[["3"]]),
        ],
    ).output


class TestAddSheet:
    def test_appends_by_default(self, three_sheet_workbook, tmp_path):
        result = sheets.add_sheet(three_sheet_workbook, "Four", output=tmp_path / "out.xlsx")
        assert result.sheets == ["One", "Two", "Three", "Four"]

    def test_inserts_at_index(self, three_sheet_workbook, tmp_path):
        result = sheets.add_sheet(
            three_sheet_workbook, "Zero", index=1, output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["Zero", "One", "Two", "Three"]

    def test_index_out_of_range_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "X", index=99, output=tmp_path / "out.xlsx")

    def test_duplicate_name_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "One", output=tmp_path / "out.xlsx")

    def test_case_only_duplicate_is_also_an_input_error(self, three_sheet_workbook, tmp_path):
        """Excel/openpyxl treat sheet names case-insensitively -- "one" next
        to "One" is a collision, not a distinct name."""
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "one", output=tmp_path / "out.xlsx")

    def test_forbidden_character_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "Bad[Name]", output=tmp_path / "out.xlsx")

    def test_over_31_characters_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "x" * 32, output=tmp_path / "out.xlsx")

    def test_empty_name_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.add_sheet(three_sheet_workbook, "", output=tmp_path / "out.xlsx")


class TestDeleteSheets:
    def test_deletes_by_position(self, three_sheet_workbook, tmp_path):
        result = sheets.delete_sheets(
            three_sheet_workbook, sheets="2", output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["One", "Three"]

    def test_deletes_by_name(self, three_sheet_workbook, tmp_path):
        result = sheets.delete_sheets(
            three_sheet_workbook, names=["Two"], output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["One", "Three"]

    def test_neither_sheets_nor_names_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.delete_sheets(three_sheet_workbook, output=tmp_path / "out.xlsx")

    def test_both_sheets_and_names_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.delete_sheets(
                three_sheet_workbook, sheets="1", names=["Two"], output=tmp_path / "out.xlsx"
            )

    def test_deleting_every_visible_sheet_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.delete_sheets(three_sheet_workbook, sheets="all", output=tmp_path / "out.xlsx")

    def test_deleting_all_but_one_succeeds(self, three_sheet_workbook, tmp_path):
        result = sheets.delete_sheets(
            three_sheet_workbook, sheets="1-2", output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["Three"]


class TestRenameSheet:
    def test_renames(self, three_sheet_workbook, tmp_path):
        result = sheets.rename_sheet(
            three_sheet_workbook, "Two", "Renamed", output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["One", "Renamed", "Three"]

    def test_unknown_old_name_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.rename_sheet(three_sheet_workbook, "Nope", "New", output=tmp_path / "out.xlsx")

    def test_renaming_to_an_existing_name_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError):
            sheets.rename_sheet(three_sheet_workbook, "Two", "Three", output=tmp_path / "out.xlsx")

    def test_renaming_to_a_case_only_duplicate_is_also_an_input_error(
        self, three_sheet_workbook, tmp_path
    ):
        with pytest.raises(InputError):
            sheets.rename_sheet(three_sheet_workbook, "Two", "three", output=tmp_path / "out.xlsx")

    def test_a_case_only_rename_of_the_same_sheet_succeeds_exactly(
        self, three_sheet_workbook, tmp_path
    ):
        """openpyxl's own title setter runs its own case-insensitive
        uniqueness check that does not exclude the sheet's own prior name,
        so `ws.title = "two"` next to the sheet's own current title "Two"
        would otherwise silently become "two1" instead of "two" -- verified
        directly against openpyxl 3.1.5. The requested name must never
        silently gain a numeric suffix."""
        result = sheets.rename_sheet(
            three_sheet_workbook, "Two", "two", output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["One", "two", "Three"]
        assert "two1" not in result.sheets

    def test_a_case_only_rename_works_even_when_the_scratch_title_is_taken(self, tmp_path):
        """The scratch title itself must avoid colliding with a real sheet
        name, case-insensitively -- a workbook that happens to have a sheet
        literally named "~rename" must not make the rename pick the wrong
        (already-occupied) scratch title."""
        out = write.create(
            tmp_path / "src.xlsx",
            sheets=[
                SheetSpec(name="Data", rows=[["x"]]),
                SheetSpec(name="~rename", rows=[["y"]]),
                SheetSpec(name="~Rename2", rows=[["z"]]),
            ],
        ).output
        result = sheets.rename_sheet(out, "Data", "data", output=tmp_path / "out.xlsx")
        assert "data" in result.sheets
        assert "~rename" in result.sheets
        assert "~Rename2" in result.sheets
        assert len(result.sheets) == 3

    def test_renaming_to_a_forbidden_character_is_an_input_error(
        self, three_sheet_workbook, tmp_path
    ):
        with pytest.raises(InputError):
            sheets.rename_sheet(
                three_sheet_workbook, "Two", "Bad:Name", output=tmp_path / "out.xlsx"
            )

    def test_refuses_when_a_formula_elsewhere_references_the_old_name(self, tmp_path):
        """openpyxl does not rewrite `=Data!A1` when `Data` is renamed --
        verified directly: after rename+save+reload the formula text is
        still `=Data!A1` even though the sheet is now named something else.
        A rename that proceeded anyway would silently leave the formula
        pointed at a sheet that no longer exists."""
        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb.active["A1"] = 42
        summary = wb.create_sheet("Summary")
        summary["A1"] = "=Data!A1"
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError, match="Data"):
            sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_refuses_when_a_defined_name_references_the_old_name(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb.active["A1"] = 42
        wb.create_sheet("Summary")
        wb.defined_names["MyName"] = DefinedName("MyName", attr_text="'Data'!$A$1")
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError, match="MyName"):
            sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_refuses_on_a_quoted_reference_to_a_name_with_spaces_and_an_apostrophe(self, tmp_path):
        """A sheet name containing a space or apostrophe can only be
        referenced in quoted form (`'It''s Data'!A1`, apostrophe doubled
        per Excel's own escaping -- verified against a real workbook), so
        the bare-name check alone would miss it."""
        wb = openpyxl.Workbook()
        wb.active.title = "It's Data"
        wb.active["A1"] = 42
        summary = wb.create_sheet("Summary")
        summary["A1"] = "='It''s Data'!A1"
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError):
            sheets.rename_sheet(source, "It's Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_proceeds_when_nothing_references_the_old_name(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Alone"
        wb.create_sheet("Other")
        source = tmp_path / "src.xlsx"
        wb.save(source)

        result = sheets.rename_sheet(source, "Alone", "Renamed", output=tmp_path / "out.xlsx")
        assert result.sheets == ["Renamed", "Other"]

    def test_a_reference_to_a_different_sheet_does_not_block_the_rename(self, tmp_path):
        """A formula referencing some *other* sheet must not false-positive
        just because it happens to run near the renamed sheet's name."""
        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        other = wb.create_sheet("Other")
        other["A1"] = "=Other!B1"
        source = tmp_path / "src.xlsx"
        wb.save(source)

        result = sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")
        assert result.sheets == ["Renamed", "Other"]

    def test_refuses_when_a_chart_series_references_the_old_name(self, tmp_path):
        """openpyxl does not rewrite a chart series' `numRef.f` when the
        sheet it points at is renamed -- verified directly."""
        from openpyxl.chart import LineChart, Reference

        wb = openpyxl.Workbook()
        data = wb.active
        data.title = "Data"
        for row in range(1, 6):
            data.cell(row=row, column=1, value=row)
        summary = wb.create_sheet("Summary")
        chart = LineChart()
        chart.add_data(Reference(data, min_col=1, min_row=1, max_row=5))
        summary.add_chart(chart, "C1")
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError):
            sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_refuses_when_a_conditional_formatting_rule_references_the_old_name(self, tmp_path):
        from openpyxl.formatting.rule import FormulaRule

        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        summary = wb.create_sheet("Summary")
        summary.conditional_formatting.add("A1:A2", FormulaRule(formula=["Data!A1>0"]))
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError):
            sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_refuses_when_a_data_validation_references_the_old_name(self, tmp_path):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        summary = wb.create_sheet("Summary")
        dv = DataValidation(type="list", formula1="'Data'!$A$1:$A$2")
        summary.add_data_validation(dv)
        dv.add("B1")
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError):
            sheets.rename_sheet(source, "Data", "Renamed", output=tmp_path / "out.xlsx")

    def test_refuses_on_the_first_endpoint_of_a_bare_3d_range(self, tmp_path):
        """`=SUM(Sheet1:Sheet3!A1)` sheet-qualifies `Sheet1` too, even
        though it sits before `:` rather than immediately before `!`."""
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        sheet3 = wb.create_sheet("Sheet3")
        sheet3["A1"] = "=SUM(Sheet1:Sheet3!A1)"
        source = tmp_path / "src.xlsx"
        wb.save(source)

        with pytest.raises(InputError):
            sheets.rename_sheet(source, "Sheet1", "Renamed", output=tmp_path / "out.xlsx")


class TestReorderSheets:
    def test_reorders(self, three_sheet_workbook, tmp_path):
        result = sheets.reorder_sheets(
            three_sheet_workbook, [3, 1, 2], output=tmp_path / "out.xlsx"
        )
        assert result.sheets == ["Three", "One", "Two"]

    def test_reorder_persists_through_a_reload(self, three_sheet_workbook, tmp_path):
        sheets.reorder_sheets(three_sheet_workbook, [3, 1, 2], output=tmp_path / "out.xlsx")
        idx = read.get_index(tmp_path / "out.xlsx")
        assert [s.name for s in idx.sheets] == ["Three", "One", "Two"]

    def test_missing_index_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError, match="missing"):
            sheets.reorder_sheets(three_sheet_workbook, [1, 2], output=tmp_path / "out.xlsx")

    def test_duplicated_index_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError, match="duplicated"):
            sheets.reorder_sheets(three_sheet_workbook, [1, 1, 2], output=tmp_path / "out.xlsx")

    def test_out_of_range_index_is_an_input_error(self, three_sheet_workbook, tmp_path):
        with pytest.raises(InputError, match="out of range"):
            sheets.reorder_sheets(three_sheet_workbook, [1, 2, 9], output=tmp_path / "out.xlsx")


class TestFidelityGuardIntegration:
    def test_add_sheet_refuses_on_at_risk_workbook(self, at_risk_workbook, tmp_path):
        with pytest.raises(LossyEditError):
            sheets.add_sheet(at_risk_workbook, "New", output=tmp_path / "out.xlsx")

    def test_allow_lossy_proceeds(self, at_risk_workbook, tmp_path):
        result = sheets.add_sheet(
            at_risk_workbook, "New", output=tmp_path / "out.xlsx", allow_lossy=True
        )
        assert "New" in result.sheets

    def test_allow_lossy_reports_what_was_dropped(self, at_risk_workbook, tmp_path):
        """Every sheet operation accepts allow_lossy, so every one owes the
        same report -- a SheetOpResult that stays silent about what an
        allow_lossy=True caller just agreed to lose would defeat the whole
        point of the flag never making the loss silent (spec section 6.2)."""
        result = sheets.add_sheet(
            at_risk_workbook, "New", output=tmp_path / "out.xlsx", allow_lossy=True
        )
        assert result.dropped

    def test_recalculation_required_reflects_the_source(self, cached_value_workbook, tmp_path):
        result = sheets.add_sheet(cached_value_workbook, "New", output=tmp_path / "out.xlsx")
        assert result.recalculation_required is True

    def test_recalculation_not_required_with_no_formulas(self, three_sheet_workbook, tmp_path):
        result = sheets.add_sheet(three_sheet_workbook, "New", output=tmp_path / "out.xlsx")
        assert result.recalculation_required is False
