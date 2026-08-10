"""rp_xlsx.ooxml -- opened()/save(), template retyping, and readability checks.

Unlike rp-docx/rp-pptx's .dotx/.potx handling, there is no read-side retyping
to test here -- openpyxl opens a `.xltx` natively (spec section 5.3). What
this module tests is the write-side workaround: `wb.template` is sticky, and
`save()`'s output-extension retyping is what fixes it.
"""

from __future__ import annotations

import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from rp_core.errors import InputError
from rp_xlsx import ooxml
from rp_xlsx.errors import InvalidXlsxError, MissingFileError


class TestCheckReadable:
    def test_missing_file(self, tmp_path):
        with pytest.raises(MissingFileError):
            ooxml.check_readable(tmp_path / "nope.xlsx")

    def test_directory_is_not_a_file(self, tmp_path):
        directory = tmp_path / "dir.xlsx"
        directory.mkdir()
        with pytest.raises(MissingFileError):
            ooxml.check_readable(directory)

    def test_xls_extension_names_the_extension(self, tmp_path):
        path = tmp_path / "workbook.xls"
        path.write_bytes(b"not really an xls")
        with pytest.raises(InvalidXlsxError, match=r"\.xls"):
            ooxml.check_readable(path)

    def test_xlsb_extension_names_the_extension(self, tmp_path):
        path = tmp_path / "workbook.xlsb"
        path.write_bytes(b"not really an xlsb")
        with pytest.raises(InvalidXlsxError, match=r"\.xlsb"):
            ooxml.check_readable(path)

    def test_legacy_extension_hint_names_soffice_conversion(self, tmp_path):
        path = tmp_path / "workbook.xls"
        path.write_bytes(b"data")
        with pytest.raises(InvalidXlsxError, match="soffice"):
            ooxml.check_readable(path)

    def test_valid_xlsx_renamed_to_xls_is_still_refused_on_extension(
        self, tmp_path, plain_workbook
    ):
        """Spec section 9, verified: the extension check happens before any
        content check, so a perfectly good file is refused on its name."""
        renamed = tmp_path / "renamed.xls"
        renamed.write_bytes(plain_workbook.read_bytes())
        with pytest.raises(InvalidXlsxError, match=r"\.xls"):
            ooxml.check_readable(renamed)

    def test_unsupported_extension(self, tmp_path):
        path = tmp_path / "workbook.txt"
        path.write_text("hi", encoding="utf-8")
        with pytest.raises(InvalidXlsxError):
            ooxml.check_readable(path)

    def test_corrupt_zip_is_mapped_to_invalid_xlsx_error(self, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"not a zip file at all")
        with pytest.raises(InvalidXlsxError):
            ooxml.check_readable(path)

    def test_valid_workbook_passes(self, plain_workbook):
        assert ooxml.check_readable(plain_workbook) == plain_workbook


class TestOpened:
    def test_opens_xlsx(self, plain_workbook):
        with ooxml.opened(plain_workbook) as wb:
            assert wb["Sheet1"]["A1"].value == "hello"

    def test_opens_xltx_natively(self, template_workbook_path):
        with ooxml.opened(template_workbook_path) as wb:
            assert wb.template is True

    def test_missing_file_raises_missing_file_error(self, tmp_path):
        with pytest.raises(MissingFileError), ooxml.opened(tmp_path / "nope.xlsx"):
            pass

    def test_keep_vba_true_for_xlsm(self, macro_workbook):
        with ooxml.opened(macro_workbook) as wb:
            assert wb.vba_archive is not None

    def test_keep_vba_false_for_plain_xlsx(self, plain_workbook):
        with ooxml.opened(plain_workbook) as wb:
            assert wb.vba_archive is None

    def test_caller_exception_propagates_unchanged(self, plain_workbook):
        """The yield is inside a bare try/finally, not try/except -- a
        caller's own error must not be reported as a corrupt file."""

        class Boom(Exception):
            pass

        with pytest.raises(Boom), ooxml.opened(plain_workbook):
            raise Boom("caller broke, not the file")

    def test_data_only_reads_cached_value(self, cached_value_workbook):
        with ooxml.opened(cached_value_workbook, data_only=True) as wb:
            assert wb["Sheet"]["A3"].value == 3

    def test_default_reads_formula_text(self, cached_value_workbook):
        with ooxml.opened(cached_value_workbook) as wb:
            assert wb["Sheet"]["A3"].value == "=SUM(A1:A2)"


class TestContentTypes:
    """Spec section 5.3, verified against openpyxl 3.1.5: the `.dotx`/`.potx`
    finding repeats here but inverted -- the problem is the *write* side."""

    def test_openpyxl_write_side_is_sticky_without_the_workaround(self, tmp_path):
        """The bug the workaround exists for. If a future openpyxl fixes
        this, this test fails and ooxml.save's retyping can be simplified
        (spec section 5.3's own instruction)."""
        template = tmp_path / "raw-template.xltx"
        wb = Workbook()
        wb.template = True
        wb.save(template)

        loaded = load_workbook(template)
        assert loaded.template is True
        out = tmp_path / "raw-out.xlsx"
        loaded.save(out)  # no retyping -- raw openpyxl behaviour
        with zipfile.ZipFile(out) as zf:
            content_types = zf.read("[Content_Types].xml").decode()
        assert ooxml.TEMPLATE_CONTENT_TYPE in content_types  # still a template -- the bug

    def test_save_retypes_template_to_workbook_on_xlsx_output(
        self, tmp_path, template_workbook_path
    ):
        with ooxml.opened(template_workbook_path) as wb:
            output = ooxml.save(wb, tmp_path / "retyped.xlsx")
        assert ooxml.content_type_of(output) == ooxml.WORKBOOK_CONTENT_TYPE
        assert load_workbook(output).template is False

    def test_save_retypes_workbook_to_template_on_xltx_output(self, tmp_path, plain_workbook):
        with ooxml.opened(plain_workbook) as wb:
            output = ooxml.save(wb, tmp_path / "retyped.xltx")
        assert ooxml.content_type_of(output) == ooxml.TEMPLATE_CONTENT_TYPE
        assert load_workbook(output).template is True

    def test_save_sets_full_calc_on_load(self, tmp_path, plain_workbook):
        with ooxml.opened(plain_workbook) as wb:
            output = ooxml.save(wb, tmp_path / "out.xlsx")
        assert load_workbook(output).calculation.fullCalcOnLoad is True

    def test_macro_workbook_keeps_vba_part_and_content_type(self, tmp_path, macro_workbook):
        with ooxml.opened(macro_workbook) as wb:
            output = ooxml.save(wb, tmp_path / "out.xlsm")
        with zipfile.ZipFile(output) as zf:
            assert "xl/vbaProject.bin" in zf.namelist()
        assert ooxml.content_type_of(output) == ooxml.MACRO_WORKBOOK_CONTENT_TYPE


class TestSaveValidatesTheOutputExtension:
    """openpyxl's own docstring: 'Excel requires the file extension to match
    but openpyxl does not enforce this.' save() does, in both directions,
    because a mismatch is not a corrupt file -- it opens fine and lies about
    what it is."""

    def test_an_unsupported_suffix_is_refused(self, tmp_path, plain_workbook):
        with ooxml.opened(plain_workbook) as wb, pytest.raises(InputError):
            ooxml.save(wb, tmp_path / "out.docx")

    def test_saving_a_macro_workbook_under_a_non_macro_suffix_is_refused(
        self, tmp_path, macro_workbook
    ):
        """Without this check, openpyxl keeps the macro-enabled content type
        (and the macros) under a name that claims otherwise -- verified by
        the raw-openpyxl probe below."""
        with ooxml.opened(macro_workbook) as wb, pytest.raises(InputError):
            ooxml.save(wb, tmp_path / "out.xlsx")

    def test_openpyxl_itself_mislabels_a_macro_workbook_saved_as_xlsx(
        self, tmp_path, macro_workbook
    ):
        """The bug the refusal above exists for, without rp-xlsx in the way."""
        wb = load_workbook(macro_workbook, keep_vba=True)
        try:
            out = tmp_path / "raw-out.xlsx"
            wb.save(out)  # no rp-xlsx validation -- raw openpyxl behaviour
        finally:
            # Same leak ooxml.opened()'s finally block guards against --
            # left open, its __del__ fires during GC and can warn about an
            # I/O operation on a file this test has already left behind.
            if wb.vba_archive is not None:
                wb.vba_archive.close()
            wb.close()
        with zipfile.ZipFile(out) as zf:
            assert "xl/vbaProject.bin" in zf.namelist()  # macros survived...
            content_types = zf.read("[Content_Types].xml").decode()
        assert ooxml.MACRO_WORKBOOK_CONTENT_TYPE in content_types  # ...under an .xlsx name

    def test_saving_a_plain_workbook_under_a_macro_suffix_is_refused(
        self, tmp_path, plain_workbook
    ):
        with ooxml.opened(plain_workbook) as wb, pytest.raises(InputError):
            ooxml.save(wb, tmp_path / "out.xlsm")


class TestPopulatedCells:
    """Every full-sheet scan in this package goes through this rather than
    ``ws.iter_rows()`` with no bounds, which walks the *declared* rectangle
    and can cost billions of cell constructions on a phantom dimension."""

    def test_returns_only_cells_that_exist(self, phantom_dimension_workbook):
        with ooxml.opened(phantom_dimension_workbook) as wb:
            cells = ooxml.populated_cells(wb.active)
        assert [(c.row, c.column) for c in cells] == [(1, 1), (1000, 5)]

    def test_row_major_order_regardless_of_write_order(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        # Written out of row-major order deliberately.
        ws.cell(row=3, column=1, value="c")
        ws.cell(row=1, column=2, value="a")
        ws.cell(row=1, column=1, value="a0")
        cells = ooxml.populated_cells(ws)
        assert [(c.row, c.column) for c in cells] == [(1, 1), (1, 2), (3, 1)]

    def test_stays_fast_at_excels_actual_row_and_column_limits(
        self, adversarial_phantom_dimension_workbook
    ):
        import time

        with ooxml.opened(adversarial_phantom_dimension_workbook) as wb:
            start = time.monotonic()
            cells = ooxml.populated_cells(wb.active)
            elapsed = time.monotonic() - start
        assert elapsed < 5, f"populated_cells took {elapsed:.1f}s"
        assert len(cells) == 2


class TestFormatOf:
    def test_reports_xlsx(self, plain_workbook):
        assert ooxml.format_of(plain_workbook) == "xlsx"

    def test_reports_xltx(self, template_workbook_path):
        assert ooxml.format_of(template_workbook_path) == "xltx"

    def test_reports_xlsm(self, macro_workbook):
        assert ooxml.format_of(macro_workbook) == "xlsm"
