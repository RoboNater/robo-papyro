"""rp_xlsx.templates -- resolution, inspection, manifests, synthesis
(spec sections 5.1, 5.2, and 11.2's required assertions).
"""

from __future__ import annotations

import os

import openpyxl
import pytest

from rp_xlsx.errors import TemplateError
from rp_xlsx.templates import (
    build_manifest,
    find_placeholders,
    inspect_template,
    list_templates,
    resolve_template,
    synthesize,
)


class TestResolveTemplate:
    def test_none_returns_none(self, monkeypatch):
        monkeypatch.delenv("RP_XLSX_TEMPLATE", raising=False)
        assert resolve_template(None) is None

    def test_none_uses_configured_default_env_var(self, monkeypatch, plain_workbook):
        monkeypatch.setenv("RP_XLSX_TEMPLATE", str(plain_workbook))
        assert resolve_template(None) == plain_workbook

    def test_existing_path_is_used_as_given(self, plain_workbook):
        assert resolve_template(plain_workbook) == plain_workbook

    def test_existing_path_as_string(self, plain_workbook):
        assert resolve_template(str(plain_workbook)) == plain_workbook

    def test_wrong_path_names_the_path_not_a_template_name(self, tmp_path):
        missing = tmp_path / "nope" / "quarterly.xltx"
        with pytest.raises(TemplateError, match=str(missing)):
            resolve_template(missing)

    def test_bare_name_resolves_against_template_dir_env(self, monkeypatch, tmp_path):
        directory = tmp_path / "templates"
        directory.mkdir()
        (directory / "quarterly.xltx").write_bytes(b"fake")
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", str(directory))
        assert resolve_template("quarterly") == directory / "quarterly.xltx"

    def test_xltx_preferred_over_xlsx_when_both_exist(self, monkeypatch, tmp_path):
        directory = tmp_path / "templates"
        directory.mkdir()
        (directory / "house.xltx").write_bytes(b"fake-template")
        (directory / "house.xlsx").write_bytes(b"fake-workbook")
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", str(directory))
        assert resolve_template("house") == directory / "house.xltx"

    def test_unresolvable_name_lists_available_templates(self, monkeypatch, tmp_path):
        directory = tmp_path / "templates"
        directory.mkdir()
        (directory / "quarterly.xltx").write_bytes(b"fake")
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", str(directory))
        with pytest.raises(TemplateError, match="quarterly"):
            resolve_template("does-not-exist")

    def test_template_dir_env_splits_on_pathsep(self, monkeypatch, tmp_path):
        first = tmp_path / "a"
        second = tmp_path / "b"
        first.mkdir()
        second.mkdir()
        (second / "quarterly.xltx").write_bytes(b"fake")
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", f"{first}{os.pathsep}{second}")
        assert resolve_template("quarterly") == second / "quarterly.xltx"


class TestInspectTemplate:
    def test_minimal_template(self, minimal_template):
        info = inspect_template(minimal_template)
        assert info.format == "xlsx"
        assert info.sheets[0].name == "Sheet1"
        assert info.placeholders == []

    def test_house_like_reports_every_sheet(self, house_like_template):
        info = inspect_template(house_like_template)
        names = [s.name for s in info.sheets]
        assert names == ["Report", "2", "Résumé Data"]

    def test_house_like_reports_hidden_state(self, house_like_template):
        info = inspect_template(house_like_template)
        by_name = {s.name: s for s in info.sheets}
        assert by_name["2"].state == "hidden"
        assert by_name["Report"].state == "visible"

    def test_house_like_placeholders(self, house_like_template):
        info = inspect_template(house_like_template)
        assert set(info.placeholders) == {"client.name", "report.date"}

    def test_house_like_defined_names(self, house_like_template):
        info = inspect_template(house_like_template)
        assert any(n.name == "ReportTitle" for n in info.defined_names)


class TestFindPlaceholders:
    def test_adjacent_prefix_keys_are_both_found(self, hostile_template):
        keys = find_placeholders(hostile_template)
        assert "client" in keys
        assert "client.name" in keys

    def test_stays_fast_at_excels_actual_row_and_column_limits(
        self, adversarial_phantom_dimension_workbook
    ):
        import time

        start = time.monotonic()
        find_placeholders(adversarial_phantom_dimension_workbook)
        elapsed = time.monotonic() - start
        assert elapsed < 5, f"find_placeholders took {elapsed:.1f}s"


class TestListTemplates:
    def test_lists_every_resolvable_template(self, monkeypatch, tmp_path):
        directory = tmp_path / "templates"
        directory.mkdir()
        openpyxl.Workbook().save(directory / "alpha.xlsx")
        openpyxl.Workbook().save(directory / "beta.xlsx")
        monkeypatch.setenv("RP_XLSX_TEMPLATE_DIR", str(directory))
        names = {info.name for info in list_templates()}
        assert names == {"alpha", "beta"}


class TestBuildManifest:
    def test_redaction_body_text_never_appears(self, tmp_path):
        """Spec section 5.2: a manifest built from a template whose body
        cells contain distinctive text must not contain that text anywhere
        in its serialized form."""
        path = tmp_path / "redact.xltx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount"])
        ws.append(["DISTINCTIVE-BODY-TEXT-12345", 1])
        wb.template = True
        wb.save(path)

        manifest = build_manifest(path)
        assert "DISTINCTIVE-BODY-TEXT-12345" not in manifest.model_dump_json()

    def test_header_is_structure_and_survives(self, house_like_template):
        manifest = build_manifest(house_like_template)
        report_shape = next(s for s in manifest.sheets if s.name == "Report")
        assert report_shape.header == ["Item", "Amount", "Share", "Due"]

    def test_placeholder_cells_are_captured(self, house_like_template):
        manifest = build_manifest(house_like_template)
        report_shape = next(s for s in manifest.sheets if s.name == "Report")
        assert report_shape.placeholder_cells["A5"] == "Client: {{ client.name }}"
        assert report_shape.placeholder_cells["A6"] == "Date: {{ report.date }}"

    def test_uniform_number_formats_are_captured(self, house_like_template):
        manifest = build_manifest(house_like_template)
        report_shape = next(s for s in manifest.sheets if s.name == "Report")
        assert report_shape.number_formats["C"] == "0.00%"
        assert report_shape.number_formats["D"] == "yyyy-mm-dd"

    def test_table_names_are_captured(self, house_like_template):
        manifest = build_manifest(house_like_template)
        report_shape = next(s for s in manifest.sheets if s.name == "Report")
        assert report_shape.table_names == ["LineItems"]

    def test_image_count_not_image_bytes(self, house_like_template):
        manifest = build_manifest(house_like_template)
        assert manifest.image_count == 1


class TestSynthesize:
    """Spec section 11.2's required round trip: inspect -> manifest ->
    synthesize -> the synthesized template's shape equals the original's."""

    def test_round_trip_sheet_names_order_and_visibility(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_info = inspect_template(synth_path)
        original_info = inspect_template(house_like_template)
        assert [s.name for s in synth_info.sheets] == [s.name for s in original_info.sheets]
        assert [s.state for s in synth_info.sheets] == [s.state for s in original_info.sheets]

    def test_round_trip_headers(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_manifest = build_manifest(synth_path)
        original_by_name = {s.name: s.header for s in manifest.sheets}
        synth_by_name = {s.name: s.header for s in synth_manifest.sheets}
        assert synth_by_name == original_by_name

    def test_round_trip_number_formats(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_manifest = build_manifest(synth_path)
        original = next(s for s in manifest.sheets if s.name == "Report").number_formats
        synth = next(s for s in synth_manifest.sheets if s.name == "Report").number_formats
        assert synth == original

    def test_round_trip_defined_names(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_info = inspect_template(synth_path)
        assert {n.name for n in synth_info.defined_names} == {
            n.name for n in manifest.defined_names
        }

    def test_round_trip_table_names(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_manifest = build_manifest(synth_path)
        original = next(s for s in manifest.sheets if s.name == "Report").table_names
        synth = next(s for s in synth_manifest.sheets if s.name == "Report").table_names
        assert synth == original

    def test_round_trip_placeholder_cells(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        synth_manifest = build_manifest(synth_path)
        original = next(s for s in manifest.sheets if s.name == "Report").placeholder_cells
        synth = next(s for s in synth_manifest.sheets if s.name == "Report").placeholder_cells
        assert synth == original

    def test_synthesized_template_is_typed_as_a_template(self, house_like_template, tmp_path):
        manifest = build_manifest(house_like_template)
        synth_path = synthesize(manifest, tmp_path / "synth.xltx")
        assert openpyxl.load_workbook(synth_path).template is True
