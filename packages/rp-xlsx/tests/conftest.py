"""Fixtures generated at test time — no binary workbooks committed (spec
section 11.1). Grows through Phase 3; this step adds only what
``test_ooxml_xlsx.py`` and ``test_fidelity_xlsx.py`` need. The three named
synthetic templates (``minimal``/``house_like``/``hostile``, spec section
11.2) land in Phase 3 step 8.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage

WORKBOOK_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
MACRO_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"


def _repack(
    source: Path,
    target: Path,
    *,
    extra: dict[str, bytes] | None = None,
    replace: dict[str, bytes] | None = None,
) -> Path:
    """Copy ``source``'s zip to ``target``, adding ``extra`` parts and
    substituting ``replace`` ones — the same primitive every injection
    fixture in this suite is built from (spec section 11.1)."""
    extra = extra or {}
    replace = replace or {}
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(target, "w") as zout:
        for item in zin.infolist():
            data = replace.get(item.filename, zin.read(item.filename))
            zout.writestr(item, data)
        for name, data in extra.items():
            zout.writestr(name, data)
    return target


@pytest.fixture
def repack_zip():
    """The raw zip-injection helper, for tests that need a bespoke fixture."""
    return _repack


@pytest.fixture
def plain_workbook(tmp_path) -> Path:
    """A minimal workbook: no at-risk parts, no formulas, one text cell."""
    path = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    wb.save(path)
    return path


@pytest.fixture
def template_workbook_path(tmp_path) -> Path:
    """A workbook saved as `.xltx` — openpyxl opens templates natively
    (spec section 5.3), so no injection is needed to produce one."""
    path = tmp_path / "template.xltx"
    wb = openpyxl.Workbook()
    wb.template = True
    ws = wb.active
    ws["A1"] = "title"
    wb.save(path)
    return path


@pytest.fixture
def formula_workbook_path(tmp_path) -> Path:
    """A workbook with a formula and no cached value — openpyxl's own
    default shape, and the baseline `cached_value_workbook` injects onto."""
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["A2"], ws["A3"] = 1, 2, "=SUM(A1:A2)"
    wb.save(path)
    return path


@pytest.fixture
def cached_value_workbook(tmp_path, formula_workbook_path) -> Path:
    """The same workbook, with A3's formula carrying an injected cached
    `<v>` — the only way to test section 6.1 (spec section 11.1 item 1): an
    openpyxl-authored workbook has no cached values to lose otherwise.
    openpyxl writes an empty `<v></v>` alongside a value-less `<f>`
    (verified against 3.1.5), which is what gets replaced."""
    path = tmp_path / "cached.xlsx"
    with zipfile.ZipFile(formula_workbook_path) as zin:
        sheet = zin.read("xl/worksheets/sheet1.xml")
    injected = sheet.replace(b"<f>SUM(A1:A2)</f><v></v>", b"<f>SUM(A1:A2)</f><v>3</v>")
    assert injected != sheet, "openpyxl's empty <v/> placeholder was not found to replace"
    _repack(formula_workbook_path, path, replace={"xl/worksheets/sheet1.xml": injected})
    return path


@pytest.fixture
def at_risk_workbook(tmp_path, plain_workbook) -> Path:
    """The six representative at-risk parts from spec section 11.1 item 2 —
    presence fixtures only. They need no valid content: the guard keys on
    part names, not on anything inside them (the `modern_comments_deck`
    doctrine `AGENTS.md` and the spec both cite)."""
    path = tmp_path / "at_risk.xlsx"
    extra = {
        "customXml/item1.xml": b"<root/>",
        "xl/ctrlProps/ctrlProp1.xml": b"<formControlPr/>",
        "xl/persons/person.xml": b"<personList/>",
        "xl/pivotCache/pivotCacheDefinition1.xml": b"<pivotCacheDefinition/>",
        "xl/slicers/slicer1.xml": b"<slicer/>",
        "xl/threadedComments/threadedComment1.xml": b"<threadedComments/>",
    }
    _repack(plain_workbook, path, extra=extra)
    return path


@pytest.fixture
def macro_workbook(tmp_path, plain_workbook) -> Path:
    """A `.xlsm` carrying an injected `xl/vbaProject.bin` and the
    macro-enabled content type (spec section 11.1 item 3), to test that
    `keep_vba` preserves it across an edit."""
    path = tmp_path / "macro.xlsm"
    with zipfile.ZipFile(plain_workbook) as zin:
        content_types = zin.read("[Content_Types].xml").decode()
    content_types = content_types.replace(WORKBOOK_CONTENT_TYPE, MACRO_WORKBOOK_CONTENT_TYPE)
    if "vbaProject" not in content_types:
        content_types = content_types.replace(
            "</Types>",
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
        )
    _repack(
        plain_workbook,
        path,
        extra={"xl/vbaProject.bin": b"FAKE-VBA-BYTES"},
        replace={"[Content_Types].xml": content_types.encode()},
    )
    return path


@pytest.fixture
def rich_workbook_path(tmp_path) -> Path:
    """One workbook exercising every read.py code path at once: formulas,
    merges, a comment, an image, a chart, an Excel table, a defined name
    (workbook- and sheet-scoped), freeze panes, autofilter, dates, a
    percentage, a boolean, and a second sheet literally named ``"2"``
    (spec section 4's disambiguation case) that is hidden."""
    path = tmp_path / "rich.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["B1"], ws["C1"] = "Name", "Amount", "Done"
    ws["A2"], ws["B2"], ws["C2"] = "alpha", 10, True
    ws["A3"], ws["B3"], ws["C3"] = "beta", 20, False
    ws["B4"] = "=SUM(B2:B3)"
    ws["D2"] = 0.25
    ws["D2"].number_format = "0.00%"
    ws["E2"] = datetime(2024, 5, 1, 12, 30)
    ws["E3"] = date(2024, 5, 1)
    ws.merge_cells("A6:B6")
    ws["A6"] = "merged"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C3"
    ws["A1"].comment = Comment("header note", "Author")

    table = Table(displayName="DataTable", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    chart = BarChart()
    chart.title = "Amounts"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=3)
    chart.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "F2")

    buf = io.BytesIO()
    PILImage.new("RGB", (12, 8), color="blue").save(buf, format="PNG")
    buf.seek(0)
    ws.add_image(Image(buf), "F10")

    wb.defined_names["Revenue"] = openpyxl.workbook.defined_name.DefinedName(
        "Revenue", attr_text="Data!$B$2:$B$3"
    )
    ws.defined_names["LocalNote"] = openpyxl.workbook.defined_name.DefinedName(
        "LocalNote", attr_text="Data!$A$1"
    )

    hidden = wb.create_sheet("2")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "second sheet"

    wb.save(path)
    return path


@pytest.fixture
def phantom_dimension_workbook(tmp_path) -> Path:
    """A sheet with one real value and a format-only cell far below it, so
    `ws.dimensions`/`max_row` claim far more than the sheet actually holds
    (spec section 9, verified in the probe note)."""
    path = tmp_path / "phantom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "only real value"
    ws["E1000"].fill = openpyxl.styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    wb.save(path)
    return path


@pytest.fixture
def empty_workbook(tmp_path) -> Path:
    """A workbook with a single, genuinely empty sheet."""
    path = tmp_path / "empty.xlsx"
    openpyxl.Workbook().save(path)
    return path


# --- the three synthetic templates (spec section 11.2) -----------------------
#
# Adversarial, not realistic -- the same doctrine rp-docx and rp-pptx use for
# their own trio. No real house template ever enters this repository (spec
# section 5.2); these exist purely to exercise resolution, inspection,
# manifest building, and synthesis.


@pytest.fixture
def minimal_template(tmp_path) -> Path:
    """One sheet, a header row, no placeholders, `.xlsx`. The happy path and
    the no-template path."""
    path = tmp_path / "minimal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount"])
    wb.save(path)
    return path


@pytest.fixture
def house_like_template(tmp_path) -> Path:
    """`.xltx`; three sheets -- one hidden and named "2" (section 4's
    disambiguation case), one with a space and a non-ASCII character in its
    name; a header row; per-column number formats including a percentage
    and a date; frozen panes; a defined name; an Excel table; `{{ }}`
    placeholders in a title block; an image on the first sheet.

    The table's own header row is deliberately made to coincide with the
    sheet's first used row (rather than sitting under a separate title
    block), and the title-block placeholders sit in a single dense column
    below the table. openpyxl does not round-trip an empty-string cell value
    (verified: it comes back ``None`` after a save/reload, unlike every
    other value including a bare space), so any layout that leaves a *gap*
    inside a manifest's header row cannot be reconstructed byte-for-byte by
    ``synthesize()``. Keeping the header row itself gap-free sidesteps that
    openpyxl limitation rather than fighting it.
    """
    path = tmp_path / "house_like.xltx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["Item", "Amount", "Share", "Due"])
    ws.append(["Alpha", 100, 0.5, datetime(2024, 5, 1)])
    ws.append(["Beta", 200, 0.75, datetime(2024, 6, 1)])
    ws["C2"].number_format = "0.00%"
    ws["C3"].number_format = "0.00%"
    ws["D2"].number_format = "yyyy-mm-dd"
    ws["D3"].number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"
    table = Table(displayName="LineItems", ref="A1:D3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    ws["A5"] = "Client: {{ client.name }}"
    ws["A6"] = "Date: {{ report.date }}"
    buf = io.BytesIO()
    PILImage.new("RGB", (8, 8), color="green").save(buf, format="PNG")
    buf.seek(0)
    ws.add_image(Image(buf), "F1")
    wb.defined_names["ReportTitle"] = openpyxl.workbook.defined_name.DefinedName(
        "ReportTitle", attr_text="Report!$A$1"
    )

    hidden = wb.create_sheet("2")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "internal"

    unicode_sheet = wb.create_sheet("Résumé Data")
    unicode_sheet["A1"] = "Header"

    wb.template = True
    wb.save(path)
    return path


@pytest.fixture
def hostile_template(tmp_path) -> Path:
    """A sheet whose name is exactly 31 characters; a cell whose value
    begins with `=` but is stored as text; a merged block; a formula with
    no cached value; a column of mixed types; a format-only cell at row
    5000 (the phantom-dimension case); a placeholder split by nothing but
    adjacent to another placeholder key that is a prefix of it
    (`{{ client }}` next to `{{ client.name }}` -- the longest-first rule)."""
    path = tmp_path / "hostile.xlsx"
    long_name = "x" * 31
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = long_name
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "literal-equals"
    literal_cell = ws["B2"]
    literal_cell.value = "=NOTAFORMULA"
    literal_cell.data_type = "s"
    ws["A3"] = "formula"
    ws["B3"] = "=SUM(C1:C2)"
    ws["A4"] = "mixed"
    ws["B4"] = 42
    ws["A5"] = "mixed"
    ws["B5"] = "not a number"
    ws.merge_cells("A6:B6")
    ws["A6"] = "merged block"
    ws["A8"] = "{{ client }} and {{ client.name }} adjacent"
    ws["E5000"].fill = openpyxl.styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    wb.save(path)
    return path
